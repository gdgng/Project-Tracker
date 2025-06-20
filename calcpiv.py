
import openpyxl
import requests
import time
import json
import hmac
import hashlib
import logging
import sys
import os
import csv
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
global btc_value
from datetime import datetime
import configparser


CONFIG_FILE = "tracker.cfg"


# Load warm storage credentials from tracker.xlsx
try:
    ws = openpyxl.load_workbook('tracker.xlsx')['Credentials']
    Warm_API_Name = ws.cell(row=2,column=2).value
    WARM_API_KEY = ws.cell(row=3, column=2).value
    WARM_API_SECRET = ws.cell(row=4, column=2).value
    WARM_API_URL = ws.cell(row=5, column=2).value
except (FileNotFoundError, KeyError, Exception) as e:
    print(f"Error opening 'tracker.xlsx': {e}")
    sys.exit()



# Warm Storage API Functions
def create_signature(ts, method, endp, body=None):
    """
    Create the HMAC SHA256 signature for authentication with the Bitvavo API.
    """
    msg = str(ts) + method + '/v2/' + endp
    if body:
        msg += json.dumps(body)
    return hmac.new(WARM_API_SECRET.encode('utf-8'), msg.encode(), hashlib.sha256).hexdigest()

def warm_exchange_req(method, endpoint, params=None):
    """
    Perform an API request to Bitvavo using the given method and endpoint.
    """
    ts = int(time.time() * 1000)  # Current timestamp in milliseconds
    headers = {f'{Warm_API_Name}-Access-Key': WARM_API_KEY,
               f'{Warm_API_Name}-Access-Timestamp': str(ts),
               f'{Warm_API_Name}-Access-Signature': create_signature(ts, method, endpoint, params)}

    try:
        resp = requests.request(method, WARM_API_URL + endpoint, headers=headers, params=params)
        resp.raise_for_status()  # Raise an exception for HTTP errors
        return resp.json()  # Return the response data in JSON format
    except requests.exceptions.RequestException as e:
        logging.error(f"Warm Storage Access API error ({endpoint}): {e}")
        return None

def get_warm_exchange_ticker(market):
    """
    Fetch the ticker price for a specific market (e.g., 'BTC-EUR') from the warm exchange.
    """
    return warm_exchange_req('GET', f"ticker/price?market={market}")

def get_warm_exchange_balance():
    """
    Fetch the balance for all available coins in warm storage.
    """
    data = warm_exchange_req('GET', "balance")
    if data:
        return {item['symbol']: {'available': float(item['available']), 'in_order': float(item['inOrder'])} for item in data}
    return None

def get_crypto_ticker(crypto):
    """
    Get the price of a cryptocurrency in EUR. Returns 1.0 for EUR itself.
    """
    if crypto == "EUR":
        return {'eur_rate': 1.0, 'updated': int(time.time())}

    data = get_warm_exchange_ticker(f"{crypto}-EUR")
    if data and 'price' in data:
        return {'eur_rate': float(data['price']), 'updated': int(time.time())}

    logging.error(f"Error: Price not found for {crypto}")
    return None

def read_cold_storage_xlsx(file_path='tracker.xlsx'):
    try:
        df = pd.read_excel(file_path, sheet_name='Cold_Storage', header=1)
        return dict(zip(df['Coin'].str.upper(), df['Amount']))
    except Exception as e:
        print(f"Error reading cold storage from '{file_path}': {e}")
        return {}



def calculate_buy_stake_sell_data(csv_filepath, real_time_prices):
    data_per_coin = defaultdict(lambda: {
        # Warm storage tracking
        'bought_amount': 0,
        'bought_cost': 0,
        'staked_amount': 0,
        'deposited_amount': 0,    # crypto coming into warm storage
        'withdrawn_amount': 0,    # crypto leaving warm storage (to cold)
        'sold_amount': 0,
        'sold_revenue': 0,        # total EUR received from sells

        # Fee tracking
        'total_fees_eur': 0,
        'total_fees_crypto': 0,

        # Calculated fields (will be computed later)
        'warm_balance': 0,
        'avg_buy_price': 0,
        'avg_sell_price': 0,
        'current_warm_value': 0,
        'total_invested_this_coin': 0
    })

    eur_invested = 0.0      # Total EUR deposited (money in)
    eur_withdrawn = 0.0     # Total EUR withdrawn (money out)
    all_rows = []

    try:
        with open(csv_filepath, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                all_rows.append(row)

                # Extract and clean data
                tx_type = row.get('Type', '').lower().strip()
                currency = row.get('Currency', '').upper().strip()
                amount_str = row.get('Amount', '0')
                quote_currency = row.get('Quote Currency', '').upper().strip()
                quote_price_str = row.get('Quote Price', '0')
                received_paid_currency = row.get('Received / Paid Currency', '').upper().strip()
                received_paid_amount_str = row.get('Received / Paid Amount', '0')
                fee_currency = row.get('Fee currency', '').upper().strip()
                fee_amount_str = row.get('Fee amount', '0')

                # Convert to float with error handling
                try:
                    amount = float(amount_str) if amount_str else 0
                    quote_price = float(quote_price_str) if quote_price_str else 0
                    received_paid_amount = float(received_paid_amount_str) if received_paid_amount_str else 0
                    fee_amount = float(fee_amount_str) if fee_amount_str else 0
                except (ValueError, TypeError):
                    print(f"Warning: Could not parse numeric values in row: {row}")
                    continue

                # Skip if no amount
                if amount == 0:
                    continue

                # Process EUR transactions (investment tracking)
                if currency == 'EUR':
                    #eur_balance = get_eur_balance()

                    if tx_type == 'deposit':
                        eur_invested += amount
                    elif tx_type == 'withdrawal':
                        eur_withdrawn += abs(amount)
                    continue  # Skip further processing for EUR

                # Skip if currency not in our tracking list
                if currency not in real_time_prices:
                    continue

                # Process crypto transactions
                if tx_type == 'buy' and quote_currency == 'EUR':
                    # Buying crypto with EUR
                    cost = amount * quote_price
                    data_per_coin[currency]['bought_amount'] += amount
                    data_per_coin[currency]['bought_cost'] += cost

                elif tx_type == 'sell' and quote_currency == 'EUR':
                    # Selling crypto for EUR
                    revenue = amount * quote_price
                    data_per_coin[currency]['sold_amount'] += abs(amount)
                    data_per_coin[currency]['sold_revenue'] += revenue

                elif tx_type == 'staking':
                    # Staking rewards received
                    data_per_coin[currency]['staked_amount'] += amount

                elif tx_type == 'deposit' and currency != 'EUR':
                    # Crypto deposited to warm storage (from external source)
                    data_per_coin[currency]['deposited_amount'] += amount

                elif tx_type == 'withdrawal' and currency != 'EUR':
                    # Crypto withdrawn from warm storage (assumed to cold storage)
                    data_per_coin[currency]['withdrawn_amount'] += abs(amount)

                # Process fees
                if fee_amount > 0 and fee_currency:
                    if fee_currency == 'EUR':
                        data_per_coin[currency]['total_fees_eur'] += fee_amount
                    else:
                        # Fee paid in crypto
                        data_per_coin[fee_currency]['total_fees_crypto'] += fee_amount

        # Calculate derived values for each coin
        result = {}
        for coin, data in sorted(data_per_coin.items()):
            # Calculate warm storage balance
            warm_balance = (data['bought_amount'] + data['staked_amount'] +
                           data['deposited_amount'] - data['withdrawn_amount'] -
                           data['sold_amount'])

            # Calculate average prices
            avg_buy_price = (data['bought_cost'] / data['bought_amount']
                            if data['bought_amount'] > 0 else 0.0)

            avg_sell_price = (data['sold_revenue'] / data['sold_amount']
                             if data['sold_amount'] > 0 else 0.0)

            # Calculate current value of warm storage
            current_price = real_time_prices.get(coin, 0)
            current_warm_value = warm_balance * current_price

            # Total invested in this specific coin (bought cost)
            total_invested_this_coin = data['bought_cost']

            result[coin] = {
                # Raw amounts
                'bought': round(data['bought_amount'], 8),
                'staked': round(data['staked_amount'], 8),
                'deposited': round(data['deposited_amount'], 8),
                'withdrawn': round(data['withdrawn_amount'], 8),
                'sold': round(data['sold_amount'], 8),

                # Calculated balances
                'warm_balance': round(warm_balance, 8),
                'cold_balance': round(data['withdrawn_amount'], 8),  # Assumes withdrawals go to cold

                # Pricing
                'avg_buy_price': round(avg_buy_price, 4),
                'avg_sell_price': round(avg_sell_price, 4),
                'current_price': round(current_price, 4),

                # Values
                'total_invested': round(total_invested_this_coin, 2),
                'current_warm_value': round(current_warm_value, 2),
                'sold_revenue': round(data['sold_revenue'], 2),

                # Fees
                'fees_eur': round(data['total_fees_eur'], 2),
                'fees_crypto': round(data['total_fees_crypto'], 8)
            }

        return result, round(eur_invested, 2), round(eur_withdrawn, 2), all_rows

    except Exception as e:
        print(f"Error reading CSV: {e}")
        return {}, 0.0, 0.0, []

def get_eur_balance():
    balances = get_warm_exchange_balance()  # Roep de bestaande functie aan

    for asset in balances:
        if asset["symbol"] == "EUR":
            print(f"Je EUR-balans: {asset['available']} EUR")
            return asset["available"]

    print("Geen EUR gevonden in je Bitvavo-account.")
    return None


def flatten_cols(cols):
    """Flattens multi-level column headers into a single tuple of strings."""
    new_cols = []
    for col in cols:
        if isinstance(col, tuple):
            new_cols.append(' '.join(str(c) for c in col if c))
        else:
            new_cols.append(str(col))
    return new_cols



def create_excel_with_pivots(filename, all_data, relevant_coins):
    try:
        if os.path.exists(filename):
            book = openpyxl.load_workbook(filename)
            sheets_to_delete = ['Raw Data', 'Pivot Table Summary', 'Pivot Table Detailed']
            for sheet_name in sheets_to_delete:
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
        else:
            book = openpyxl.Workbook()

        # Write Raw Data
        df_raw = pd.DataFrame(all_data)
        ws_raw = book.create_sheet('Raw Data')
        ws_raw.append(list(df_raw.columns))
        for row in pd.DataFrame.to_numpy(df_raw):
            ws_raw.append(list(row))

        if not df_raw.empty:
            # Convert numeric columns properly
            for col in ['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount']:
                df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce')

            # Normalize coin names to uppercase
            df_raw['Currency'] = df_raw['Currency'].str.upper()

            # Filter for relevant coins and EUR
            df_filtered = df_raw[df_raw['Currency'].isin(relevant_coins) | (df_raw['Currency'] == 'EUR')]

            if not df_filtered.empty:
                # Create Summary Pivot Table
                pivot_summary = pd.pivot_table(
                    df_filtered,
                    values=['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount'],
                    index='Currency',
                    columns='Type',
                    aggfunc='sum',
                    margins=True,
                    margins_name='Total'
                )
                ws_summary = book.create_sheet('Pivot Table Summary')

                # Write the first level of headers
                first_level_summary_headers = [pivot_summary.index.name] + list(pivot_summary.columns.levels[0])
                ws_summary.append(first_level_summary_headers)

                flat_cols_summary = [pivot_summary.index.name] + flatten_cols(pivot_summary.columns)
                ws_summary.append(flat_cols_summary)

                # Merge and center the first-level headers for Summary
                header_row_summary_level_two = ws_summary[2]
                start_col_summary = 2
                for main_header in ['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount']:
                    start_merge = -1
                    end_merge = -1
                    for i, cell in enumerate(header_row_summary_level_two[1:], start=2):
                        if cell.value and cell.value.startswith(main_header):
                            if start_merge == -1:
                                start_merge = i
                            end_merge = i
                        elif start_merge != -1 and not cell.value.startswith(main_header):
                            break

                    if start_merge != -1 and end_merge != -1:
                        start_letter = get_column_letter(start_merge)
                        end_letter = get_column_letter(end_merge)
                        ws_summary.merge_cells(f'{start_letter}1:{end_letter}1')
                        ws_summary[f'{start_letter}1'].alignment = Alignment(horizontal='center')

                # Write data rows
                for index, row in pivot_summary.iterrows():
                    ws_summary.append([index] + list(row))

                # Create Detailed Pivot Table
                pivot_detailed = pd.pivot_table(
                    df_filtered,
                    values=['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount'],
                    index=['Timezone', 'Date', 'Time', 'Type', 'Currency'],
                    columns=['Quote Currency', 'Received / Paid Currency', 'Fee currency'],
                    aggfunc='sum',
                    fill_value=0
                )
                ws_detailed = book.create_sheet('Pivot Table Detailed')
                index_names = list(pivot_detailed.index.names)

                # Write the first level of headers for detailed table
                first_level_detailed_headers = index_names + list(pivot_detailed.columns.levels[0])
                ws_detailed.append(first_level_detailed_headers)

                flat_cols_detailed = index_names + flatten_cols(pivot_detailed.columns)
                ws_detailed.append(flat_cols_detailed)

                # Merge and center the first-level headers for Detailed
                header_row_detailed_level_two = ws_detailed[2]
                start_col_detailed = len(index_names) + 1
                for main_header in ['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount']:
                    start_merge_detailed = -1
                    end_merge_detailed = -1
                    for i, cell in enumerate(header_row_detailed_level_two[len(index_names):], start=len(index_names) + 1):
                        if cell.value and cell.value.startswith(main_header):
                            if start_merge_detailed == -1:
                                start_merge_detailed = i
                            end_merge_detailed = i
                        elif start_merge_detailed != -1 and not cell.value.startswith(main_header):
                            break

                    if start_merge_detailed != -1 and end_merge_detailed != -1:
                        start_letter_detailed = get_column_letter(start_merge_detailed)
                        end_letter_detailed = get_column_letter(end_merge_detailed)
                        ws_detailed.merge_cells(f'{start_letter_detailed}1:{end_letter_detailed}1')
                        ws_detailed[f'{start_letter_detailed}1'].alignment = Alignment(horizontal='center')

                # Write data rows for detailed table
                for index, row in pivot_detailed.iterrows():
                    ws_detailed.append(list(index) + list(row))

        try:
            book.save(filename)
            print(f"Excel file '{filename}' updated successfully.")
        except PermissionError:
            messagebox.showerror("Permission Error", f"Could not save '{filename}'. The file might be open in another application.")

    except Exception as e:
        print(f"Error updating Excel file with pivot tables: {e}")

# ... (na de create_excel_with_pivots functie) ...

def write_treeview_to_excel(filename, sheet_name, treeview_widget, columns_tuple,
            eur_invested, eur_withdrawn, net_invested,
            total_warm_value, total_cold_value, current_value):
    """
    Schrijft de data van een Tkinter Treeview weg naar een gespecificeerd Excel-werkblad.
    Het werkblad wordt eerst geleegd als het bestaat.
    De eerste rij bevat de datum en tijd van generatie.
    De tweede rij bevat de kolomkoppen.
    """
    try:
        # Probeer het bestaande werkboek te laden.
        # Het script gaat ervan uit dat tracker.xlsx bestaat vanwege het laden van credentials.
        try:
            book = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"Het bestand '{filename}' kon niet gevonden worden om '{sheet_name}' op te slaan. Zorg dat het bestand bestaat.")
            print(f"Fout: '{filename}' niet gevonden bij poging tot opslaan van '{sheet_name}'.")
            return # Stop de functie als het bestand niet bestaat

        # Verwijder het werkblad als het bestaat om het "leeg te maken"
        if sheet_name in book.sheetnames:
            del book[sheet_name]

        # Maak het nieuwe (lege) werkblad aan
        ws_history = book.create_sheet(sheet_name)

        # Schrijf de huidige datum en tijd in de eerste cel van de eerste rij
        current_datetime_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_history.append([f"Data as shown on screen, generated on: {current_datetime_str}"])

        # Schrijf de kolomkoppen (deze komen op de tweede rij)
        ws_history.append(list(columns_tuple))

        # Verzamel en schrijf de data uit de treeview
        data_written_to_sheet = False
        for item_id in treeview_widget.get_children():
            row_values = treeview_widget.item(item_id, 'values')
            # Converteer alle waarden naar string om typeproblemen met openpyxl te voorkomen
            # en om de weergave "zoals op het scherm" te behouden.
            ws_history.append([str(val) for val in row_values])
            data_written_to_sheet = True

        if not data_written_to_sheet and treeview_widget.get_children():
            # Dit gebeurt als de enige rij de "No relevant transactions found." melding is.
            # De .item(item_id, 'values') geeft dan een tuple met één element.
            # De append hierboven zou dat correct moeten verwerken.
            # Als er echt helemaal niks in de treeview staat (ook niet de melding),
            # dan wordt er niets geschreven behalve de datum en headers.
            pass # De headers zijn al geschreven.

        # Kolombreedtes instellen voor betere leesbaarheid
        # Headers staan op rij 2, data begint op rij 3
        for col_idx, column_title in enumerate(columns_tuple, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            # Check lengte van de header (op rij 2)
            header_cell_value = ws_history.cell(row=2, column=col_idx).value
            if header_cell_value:
                max_length = max(max_length, len(str(header_cell_value)))

            # Check lengte van data cellen (vanaf rij 3)
            for row_idx in range(3, ws_history.max_row + 1):
                cell_value = ws_history.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Voeg wat padding toe en stel de breedte in (met een maximum)
            adjusted_width = min(max_length + 3, 50) # Max breedte 50
            if adjusted_width > 0: # Zorg ervoor dat de breedte positief is
                 ws_history.column_dimensions[column_letter].width = adjusted_width
        # Get the summary at the end
        ws=ws_history
        last_row = ws.max_row
        print(last_row)
        ws.cell(row=last_row+2, column=2, value="Euro Invested")
        ws.cell(row=last_row+3, column=2, value=eur_invested)
        ws.cell(row=last_row+2, column=3, value="Euro Withdrawn")
        ws.cell(row=last_row+3, column=3, value=eur_withdrawn)
        ws.cell(row=last_row+2, column=4, value="Current Net Invest")
        ws.cell(row=last_row+3, column=4, value=net_invested)
        ws.cell(row=last_row+2, column=6, value="Warm Value")
        ws.cell(row=last_row+3, column=6, value=total_warm_value)
        ws.cell(row=last_row+2, column=7, value="Cold Value")
        ws.cell(row=last_row+3, column=7, value=round(total_cold_value,2))
        ws.cell(row=last_row+2, column=8, value="Total Value Value")
        ws.cell(row=last_row+3, column=8, value=round(current_value,2))
        ws.cell(row=last_row+5, column=2, value="P/L Value")
        pl_value = current_value-net_invested
        ws.cell(row=last_row+6, column=2, value=round(pl_value,2))
        ws.cell(row=last_row+5, column=3, value="P/L %")
        pl_percentage=(((pl_value) / net_invested) * 100) - 100

        ws.cell(row=last_row+6, column=3, value=round(pl_percentage,2))

        book.save(filename)
        print(f"Treeview data succesful writen to '{filename}' in sheet '{sheet_name}'.")

    except PermissionError:
        messagebox.showerror("Permission Error", f"Could not save'{filename}' for '{sheet_name}'. Is the file in use???")
    except Exception as e:
        print(f"Fout bij het wegschrijven van treeview data naar Excel ('{sheet_name}'): {e}")
        messagebox.showerror("Excel Error", f"Fout bij het wegschrijven van '{sheet_name}': {e}")


def browse_file():
    filepath = filedialog.askopenfilename(
        title="Select Bitvavo CSV file",
        filetypes=[("CSV files", "*.csv")],
        initialdir=os.getcwd()
    )
    if filepath:
        # Reading cold storage coins from the tracker.xlsx
        cold_storage_data = read_cold_storage_xlsx()

        # Reading warm storage coins
        warm_storage_data = get_warm_exchange_balance()

        # Combine cold and warm storage coins into a list of relevant coins
        relevant_coins = list(set(cold_storage_data.keys()).union(set(warm_storage_data.keys())))

        # Fetch real-time prices for relevant coins
        real_time_prices = {}
        for coin in relevant_coins:
            if coin == 'EUR':
                real_time_prices[coin] = 1.0
            else:
                ticker_data = get_warm_exchange_ticker(f"{coin}-EUR")
                if ticker_data and 'price' in ticker_data:
                    real_time_prices[coin] = float(ticker_data['price'])
                else:
                    real_time_prices[coin] = 0  # Default to 0 if price not found
                    print(f"Warning: Could not fetch price for {coin}. Setting price to 0.")

        results, eur_invested, eur_withdrawn, all_csv_data = calculate_buy_stake_sell_data(filepath, real_time_prices)

        # Clear existing data
        for row in treeview.get_children():
            treeview.delete(row)

        if results:
            total_warm_value = 0
            total_cold_value = 0
            total_invested_all = 0

            for coin, data in results.items():
                # Get cold storage amount for this coin from tracker.xlsx
                cold_amount = cold_storage_data.get(coin, 0)
                cold_value = cold_amount * data['current_price']

                # Only show coins with significant amounts
                if (data['warm_balance'] >= 0.00000001 or cold_amount >= 0.00000001):
                    treeview.insert('', 'end', values=(
                        coin,
                        f"{data['warm_balance']:.8f}",
                        f"{cold_amount:.8f}",
                        f"€{data['avg_buy_price']:.4f}" if data['avg_buy_price'] > 0 else "N/A",
                        f"€{data['avg_sell_price']:.4f}" if data['avg_sell_price'] > 0 else "N/A",
                        f"€{data['current_warm_value']:.2f}",
                        f"€{cold_value:.2f}",
                        f"€{data['total_invested']:.2f}",
                        f"€{data['fees_eur']:.2f}"
                    ))

                    total_warm_value += data['current_warm_value']
                    total_cold_value += cold_value
                    total_invested_all += data['total_invested']

            # Update all labels
            net_invested = eur_invested - eur_withdrawn
            eur_in_label.config(text=f"Total EUR In: €{eur_invested:.2f}")
            eur_out_label.config(text=f"Total EUR Out: €{eur_withdrawn:.2f}")
            total_invest_label.config(text=f"Net Invested: €{net_invested:.2f}")
            warm_value_label.config(text=f"Warm Value: €{total_warm_value:.2f}")
            cold_value_label.config(text=f"Cold Value: €{total_cold_value:.2f}")
            total_value_label.config(text=f"Total Value: €{(total_warm_value + total_cold_value):.2f}")

            finished_button.config(state=tk.NORMAL)
        else:
            treeview.insert('', 'end', values=("No relevant transactions found.",))
            eur_in_label.config(text="Total EUR In: €0.00")
            eur_out_label.config(text="Total EUR Out: €0.00")
            total_invest_label.config(text="Net Invested: €0.00")
            warm_value_label.config(text="Warm Value: €0.00")
            cold_value_label.config(text="Cold Value: €0.00")
            total_value_label.config(text="Total Value: €0.00")
            finished_button.config(state=tk.DISABLED)

        current_value=total_warm_value+total_cold_value
        if par_write_csv is True:
            write_treeview_to_excel("tracker.xlsx", "CSV_History", treeview, columns,
                            eur_invested, eur_withdrawn, net_invested,
                            total_warm_value, total_cold_value, current_value)
            create_excel_with_pivots("tracker.xlsx", all_csv_data, relevant_coins if results else [])

def load_app_settings():
    """
    Load all configuration settings into a flat dictionary for the main application.
    """
    config = configparser.ConfigParser()

    # Default settings - ensures all keys exist
    app_settings = {
        # Refresh rates
        'refresh_main': 3,
        'refresh_warm': 3,
        'refresh_cold': 3,
        'refresh_total': 3,

        # Write data options
        'write_warm': False,
        'write_cold': False,
        'write_total': True,
        'write_csv': True,

        # Optional URLs and Names
        'url1': '',
        'Name1': '',
        'url2': '',
        'Name2': '',
        'url3': '',
        'Name3': '',

        # Miscellaneous options
        'debug_mode': False,
        'dark_mode': True,
        'notifications': False,
        'cold_storage_available': False
    }

    try:
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)

            # Load RefreshRate section
            if config.has_section('RefreshRate'):
                app_settings['refresh_main'] = config.getint('RefreshRate', 'Main', fallback=3)
                app_settings['refresh_warm'] = config.getint('RefreshRate', 'Warm', fallback=3)
                app_settings['refresh_cold'] = config.getint('RefreshRate', 'Cold', fallback=3)
                app_settings['refresh_total'] = config.getint('RefreshRate', 'Total', fallback=3)

            # Load WriteData section
            if config.has_section('WriteData'):
                app_settings['write_warm'] = config.getboolean('WriteData', 'Warm', fallback=False)
                app_settings['write_cold'] = config.getboolean('WriteData', 'Cold', fallback=False)
                app_settings['write_total'] = config.getboolean('WriteData', 'Total', fallback=True)
                app_settings['write_csv'] = config.getboolean('WriteData', 'CSV', fallback=True)

            # Load OptionalURL section
            if config.has_section('OptionalURL'):
                app_settings['url1'] = config.get('OptionalURL', 'URL1', fallback='')
                app_settings['Name1'] = config.get('OptionalURL', 'Name1', fallback='')
                app_settings['url2'] = config.get('OptionalURL', 'URL2', fallback='')
                app_settings['Name2'] = config.get('OptionalURL', 'Name2', fallback='')
                app_settings['url3'] = config.get('OptionalURL', 'URL3', fallback='')
                app_settings['Name3'] = config.get('OptionalURL', 'Name3', fallback='')

            # Load Miscellaneous section
            if config.has_section('Miscellaneous'):
                app_settings['debug_mode'] = config.getboolean('Miscellaneous', 'DebugMode', fallback=False)
                app_settings['dark_mode'] = config.getboolean('Miscellaneous', 'darkmod', fallback=True)
                app_settings['notifications'] = config.getboolean('Miscellaneous', 'Notifications', fallback=False)
                app_settings['cold_storage_available'] = config.getboolean('Miscellaneous', 'Cold Storage Available', fallback=False)

    except Exception as e:
        print(f"Error loading configuration: {e}")
        print("Using default settings...")

    return app_settings

def exit_program():

    root.destroy()


# GUI setup
root = tk.Tk()
root.title("Bitvavo CSV Analyzer")
root.geometry("1100x650") # Increased height to accommodate the button
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

tk.Label(root, text="Select Bitvavo CSV File", font=("Helvetica", 14)).pack(pady=10)
tk.Button(root, text="Browse...", command=browse_file, width=20).pack(pady=5)

app_settings = load_app_settings()

if app_settings:


    print(f"   - Main Refresh Rate: {app_settings['refresh_main']} seconds")
    darkmod = app_settings['dark_mode']

    par_write_csv = app_settings.get('write_csv')
    print(par_write_csv)

columns = ('Coin', 'Warm Amount', 'Cold Amount', 'Avg Buy €', 'Avg Sell €',
          'Warm Value €', 'Cold Value €', 'Invested €', 'Fees €')
treeview = ttk.Treeview(root, columns=columns, show='headings')

for col in columns:
    treeview.heading(col, text=col)
    if col in ['Warm Amount', 'Cold Amount']:
        treeview.column(col, minwidth=120, width=120, stretch=True)
    elif 'Value' in col or 'Invested' in col or 'Fees' in col:
        treeview.column(col, minwidth=100, width=100, stretch=True)
    else:
        treeview.column(col, minwidth=80, width=100, stretch=True)



treeview.pack(padx=10, pady=10, fill="both", expand=True)

scrollbar_y = tk.Scrollbar(root, orient="vertical", command=treeview.yview)
scrollbar_y.pack(side="right", fill="y")
treeview.configure(yscrollcommand=scrollbar_y.set)

scrollbar_x = tk.Scrollbar(root, orient="horizontal", command=treeview.xview)
scrollbar_x.pack(side="bottom", fill="x")
treeview.configure(xscrollcommand=scrollbar_x.set)

totals_frame = tk.Frame(root)
totals_frame.pack(pady=10)




totals_frame = tk.Frame(root)
totals_frame.pack(pady=10)

# First row of labels
first_row = tk.Frame(totals_frame)
first_row.pack()

eur_in_label = tk.Label(first_row, text="Total EUR In: €0.00", font=("Helvetica", 10))
eur_in_label.pack(side="left", padx=10)

eur_out_label = tk.Label(first_row, text="Total EUR Out: €0.00", font=("Helvetica", 10))
eur_out_label.pack(side="left", padx=10)

total_invest_label = tk.Label(first_row, text="Net Invested: €0.00", font=("Helvetica", 10))
total_invest_label.pack(side="left", padx=10)

# Second row of labels
second_row = tk.Frame(totals_frame)
second_row.pack()

warm_value_label = tk.Label(second_row, text="Warm Value: €0.00", font=("Helvetica", 10), fg="orange")
warm_value_label.pack(side="left", padx=10)

cold_value_label = tk.Label(second_row, text="Cold Value: €0.00", font=("Helvetica", 10), fg="blue")
cold_value_label.pack(side="left", padx=10)

total_value_label = tk.Label(second_row, text="Total Value: €0.00", font=("Helvetica", 12, "bold"), fg="green")
total_value_label.pack(side="left", padx=10)





#eur_in_label = tk.Label(totals_frame, text="Total EUR In: € 0.00", font=("Helvetica", 12))
#eur_in_label.pack(side="left", padx=10)

#eur_out_label = tk.Label(totals_frame, text="Total EUR Out: € 0.00", font=("Helvetica", 12))
#eur_out_label.pack(side="left", padx=10)

#total_invest_label = tk.Label(totals_frame, text="Total Invest: € 0.00", font=("Helvetica", 12))
#total_invest_label.pack(side="left", padx=10)

finished_button = tk.Button(root, text="Finished", command=exit_program, bg="lightgray", fg="forestgreen", font=("Helvetica", 12), state=tk.DISABLED)
finished_button.pack(pady=0, padx=10, anchor="e") # Anchor to the right

root.mainloop()
