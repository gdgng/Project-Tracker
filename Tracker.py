# -------------------------------------------------------------------------------------------------
# File: tracker.py
# GDGNG
# 12-04-2025 V0.9
# 14-04-2025 V0.95
# 17-04-2025 V1.0
# 12-05-2025 V1.01 Changes in update warm,cold & total assets (no update_gui when shown),
#            visual changes to Fear and Greed
# 21-05-2025 V1.05 Changes in correctly handling
#            update screens for Warm, Cold and Total. No label errors. Write Total after exit to tracker
# 03-06-2025 EXTRA Graphical Adjustment (ICON's for Coins, ICON labels for screens, rework on cold and warm storage)
#--------------------------------------------------------------------------------------------------
# Bitcoin_tracker EUR/USD Value. Gets the EUR value from an exchange, site scraping for the current
# dollar value.
# Main Screen shows the current BITCOIN price in EUR and USD; changes every 5 seconds.
# Shows the ATH in EUR/US and Current rate EUR/USD
#
# Buttons:
# Shows Warm Storage, Cold Storage, Stocks, and Total Assets
# Gets WARM balance from Ban exchange (Token, Amount, Inorder, calculated(total) and Current_coin_price
# Gets Cold balance: the Coins and Amount from the Excel sheet tracker.xls.
# Also, the Key and Secret key from your WARM storage should be in this sheet (only read!).
# Gets the trading stock and amount from the bank api
#---------------------------------------------------------------------------------------------------
# Had to learn Python for this, and with AI help, it was fun (Thanks Co-pilot, Gemini, ChatGPT)
# Debugging with AI can be a hassle. But guiding AI in the right direction helps. Trying to correct
# mistakes AI still makes, this needs another (human) way of thinking to resolve the problem.
# --------------------------------------------------------------------------------------------------
import tkinter as tk
from tkinter import ttk
from tkinter import Menu
from tkinter import messagebox
from tkinter import Button, PhotoImage
from PIL import Image, ImageTk  # For image resizing
import tkinter.scrolledtext as tkscroll
from datetime import date, datetime
import requests
from bs4 import BeautifulSoup
import time
import threading
import logging
import sys
import hmac
import hashlib
import openpyxl
import subprocess
import os
import json
import webview
import markdown
import configparser
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill
from openpyxl.utils import get_column_letter
from functools import partial
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg#
#from calcpiv import load_csv_calculate
import logging
from functools import wraps
#logging.basicConfig(level=logging.DEBUG)

# --- Global Variables (Module Level Initialization) ---
# These variables are declared here to ensure they exist in the global scope
# before any function might try to access them. They are initialized to None
# or default values. Functions like show_total_assets will then manage them.
global bg_color, fg_color, fg_cold, fg_cyan, fg_day, fg_ani, darkmod, fg_tot_assets
global fg_tot_storage, debugmode
global par_refresh_main, par_debug_mode, par_demo_mode

os.environ['PYTHONIOENCODING'] = 'utf-8'

CONFIG_FILE = "tracker.cfg"
par_debug_mode = False
is_tracker_active = False
darkmod = False
fg_cold=""
fg_cyan=""
fg_color=""
bg_color=""
fg_day=""
fg_ani=""
fg_tot_storage=""
updater_job_total = None
status_label_total = None
status_label_total = None
btc_label = None
back_button = None
current_warm_data = []
par_demo_mode=False

# Other globals needed for calculations that might be updated elsewhere
total_stocks = 0.0
T_EUR_I = 0.0
T_EUR_O = 0.0
T_INVST = 0.0
T_PL = 0.0
pl_percentage = 0.0
total_stock_value = 0.0

# Tkinter StringVars for updating labels - also declared globally
warm_value_var = None
cold_value_var = None
total_assets_value_var = None
total_perc_var = None
total_crypto_text_var = None
total_pl_var = None
# GLOBAL VARIABLES (declared but initially set to None or default values)
# These are the variables that need to be accessible and modified by multiple functions
# within your application, particularly the screen-specific ones.
# They are set to None here so Python knows they're meant to be global,
# but their actual initialization happens inside show_total_assets.
is_tracker_active = False
updater_job_total = None
status_label_total = None
btc_label = None
back_button = None


# Other globals needed for calculations that might be updated elsewhere
total_stocks = 0.0
T_EUR_I = 0.0
T_EUR_O = 0.0
T_INVST = 0.0
T_PL = 0.0
pl_percentage = 0.0

# Tkinter StringVars for updating labels - also declared globally
warm_value_var = None
cold_value_var = None
total_assets_value_var = None
total_perc_var = None
total_crypto_text_var = None
total_pl_var = None

logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("PIL.PngImagePlugin").setLevel(logging.WARNING)
logging.getLogger("root").setLevel(logging.WARNING)

def debug_log(func):
    # Haal een specifieke logger op voor deze functie in de "debug" namespace
    log = logging.getLogger(f"debug.{func.__name__}")
    # Het level wordt al door basicConfig ingesteld, dus dit is niet per se nodig
    # log.setLevel(logging.DEBUG)

    @wraps(func)
    def wrapper(*args, **kwargs):
        log.debug(f"→ {func.__name__} called with args: {args}, kwargs: {kwargs}")
        result = func(*args, **kwargs)
        log.debug(f"← {func.__name__} returned: {result}")
        return result
    return wrapper



os.system('cls')
print('Initializing......')
stocks=0
today = date.today()
try:
    ws = openpyxl.load_workbook('tracker.xlsx')['Credentials']
    Warm_API_Name = ws.cell(row=2,column=2).value
    WARM_API_KEY = ws.cell(row=3, column=2).value
    WARM_API_SECRET = ws.cell(row=4, column=2).value
    WARM_API_URL = ws.cell(row=5, column=2).value
except (FileNotFoundError, KeyError, Exception) as e:
    print(f"Error opening: 'tracker.xlsx': {e}")
    sys.exit()

previous_prices = {}
after_id = None
selected_coin = None
available_coins=[]
#available_coins = ["BTC", "ETH", "SOL", "ADA", "POLS"]
#coin_symbols = {"BTC": "₿", "ETH": "Ξ", "SOL": "◎", "ADA": "₳", "POLS": ""}
coin_symbols = {}
stop_event = threading.Event()
balances = {}
is_tracker_active = True
menubar = None
main_widgets = {} # Initialise main_widgets as an empty dictionary
#
# Only one instance allowed for Aggr
#
aggr_window_instance = None
total_stocks = 0

#@debug_log
def create_signature(ts, method, endpoint, body=None):
    url_path = '/v2/' + endpoint  # Ensure this is correct
    msg = str(ts) + method + url_path
    if body:
        msg += json.dumps(body)
    signature = hmac.new(WARM_API_SECRET.encode('utf-8'), msg.encode(), hashlib.sha256).hexdigest()
    logging.debug(f"SIGNATURE INPUT ({endpoint}): {msg}")
    logging.debug(f"GENERATED SIGNATURE ({endpoint}): {signature}")
    return signature

def warm_exchange_req(method, endpoint, params=None, retries=3):
    ts = int(time.time() * 1000)
    headers = {f'{Warm_API_Name}-Access-Key': WARM_API_KEY,
               f'{Warm_API_Name}-Access-Timestamp': str(ts),
               f'{Warm_API_Name}-Access-Signature': create_signature(ts, method, endpoint, params),
               f'{Warm_API_Name}-Access-Window': '10000'
              }
    try:
        full_url = WARM_API_URL + endpoint
        #logging.debug(f"Request URL: {full_url}")
        #logging.debug(f"Request Method: {method}")
        #logging.debug(f"Request Headers: {headers}")
        #logging.debug(f"Request Params: {params}")
        resp = requests.request(method, full_url, headers=headers, params=params)
        #logging.debug(f"warm_exchange_req * successfull, Response Headers: {resp.headers}")
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as e:
        logging.error(f"Warm Storage Access API error ({endpoint}): {e}")
        if resp is not None and resp.status_code == 403 and retries > 0:
            #logging.warning(f"Received 403, retrying in 5 seconds... (Retries left: {retries})")
            time.sleep(5)
            return warm_exchange_req(method, endpoint, params, retries - 1)
        return None

def get_warm24_exchange_ticker(market):

    return warm_exchange_req('GET', f"ticker/24h?market={market}")

def get_warm_exchange_ticker(market):

    return warm_exchange_req('GET', f"ticker/price?market={market}")


def get_coin_exchange_ticker(market):
    return warm_exchange_req('GET', f"ticker/price?market={market}")

def get_coin_change(market):
    # Specific BITVAVO to get the 24h changes
    if market=='EUR':
        last_change=0
        return last_change

    market24h=f"{market}-EUR"
    ticker24h_data = get_warm24_exchange_ticker(market24h)

    if (ticker24h_data and 'open' in ticker24h_data and ticker24h_data['open']
        is not None and 'last' in ticker24h_data and ticker24h_data['last']
        is not None):
        try:
            open_price = float(ticker24h_data['open'])
            last_price = float(ticker24h_data['last'])
            last_change = ((last_price - open_price) / open_price) * 100
        except ValueError:
            last_change = None
            print("Could not format open/last price to float")
    else:
        last_change = 0.1
        print("Coin --> ", market)
        print("Open or last price not available in ticker24h response")
    return last_change

def get_warm_exchange_balance():
    global par_demo_mode

    if par_demo_mode:
        return {
            'OP':   {'available': 1000.0, 'in_order': 0.0},
            'EUR':  {'available': 1000.0, 'in_order': 0.0},
            'MANA': {'available': 1000.0, 'in_order': 0.0},
            'HYPE': {'available': 1000.0, 'in_order': 1.0},
#            'VET':  {'available': 10.0, 'in_order': 0.0},
            'DOGE': {'available': 1000.0, 'in_order': 0.0},
#            'ATOM': {'available': 10.0, 'in_order': 0.0},
#            'ETH':  {'available': 1.0,  'in_order': 1.0},
            'ICX':  {'available': 1000.0, 'in_order': 0.0}
        }

    data = warm_exchange_req('GET', "balance")

    if data:
        return {
            item['symbol']: {
                'available': float(item['available']),
                'in_order': float(item['inOrder'])
            }
            for item in data
        }

    return None


def get_crypto_ticker(crypto):
    if crypto == "EUR":
        # Special case: EUR itself has a fixed rate of 1.0 to EUR
        return {'eur_rate': 1.0, 'updated': int(time.time())}

    data = get_warm_exchange_ticker(f"{crypto}-EUR")

    if data and 'price' in data:
        return {'eur_rate': float(data['price']), 'updated': int(time.time())}



    logging.error(f"Error: Price not found for {crypto}")
    return None




def scrape_eur_usd():
    try:
        soup = BeautifulSoup(requests.get("https://www.wisselkoers.nl/dollar").content, 'html.parser')
        el = soup.find('span', class_='euro-unit')
        if el:
            try:
                return float(el.text.strip().split()[0].replace(',', '.'))
            except ValueError:
                logging.error(f"Failed to parse EUR/USD rate: {el.text}")
    except Exception as e:
        logging.error(f"Error scraping EUR/USD rate: {e}")
    return None

def load_app_settings():
    """
    Load all configuration settings into a flat dictionary for the main application.
    """
    config = configparser.ConfigParser()

    # Default settings - ensures all keys exist
    app_settings = {
        # Refresh rates
        'refresh_main': 3,
        'refresh_warm': 3,
        'refresh_cold': 3,
        'refresh_total': 3,

        # Write data options
        'write_warm': False,
        'write_cold': False,
        'write_total': True,
        'write_csv': True,

        # Optional URLs and Names
        'url1': '',
        'Name1': '',
        'url2': '',
        'Name2': '',
        'url3': '',
        'Name3': '',

        # Miscellaneous options
        'debug_mode': False,
        'dark_mode': True,
        'demo_mode': False,
        'cold_storage_available': False
    }

    try:
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)

            # Load RefreshRate section
            if config.has_section('RefreshRate'):
                app_settings['refresh_main'] = config.getint('RefreshRate', 'Main', fallback=3)
                app_settings['refresh_warm'] = config.getint('RefreshRate', 'Warm', fallback=3)
                app_settings['refresh_cold'] = config.getint('RefreshRate', 'Cold', fallback=3)
                app_settings['refresh_total'] = config.getint('RefreshRate', 'Total', fallback=3)

            # Load WriteData section
            if config.has_section('WriteData'):
                app_settings['write_warm'] = config.getboolean('WriteData', 'Warm', fallback=False)
                app_settings['write_cold'] = config.getboolean('WriteData', 'Cold', fallback=False)
                app_settings['write_total'] = config.getboolean('WriteData', 'Total', fallback=True)
                app_settings['write_csv'] = config.getboolean('WriteData', 'CSV', fallback=True)

            # Load OptionalURL section
            if config.has_section('OptionalURL'):
                app_settings['url1'] = config.get('OptionalURL', 'URL1', fallback='')
                app_settings['Name1'] = config.get('OptionalURL', 'Name1', fallback='')
                app_settings['url2'] = config.get('OptionalURL', 'URL2', fallback='')
                app_settings['Name2'] = config.get('OptionalURL', 'Name2', fallback='')
                app_settings['url3'] = config.get('OptionalURL', 'URL3', fallback='')
                app_settings['Name3'] = config.get('OptionalURL', 'Name3', fallback='')

            # Load Miscellaneous section
            if config.has_section('Miscellaneous'):
                app_settings['debug_mode'] = config.getboolean('Miscellaneous', 'DebugMode', fallback=False)
                app_settings['dark_mode'] = config.getboolean('Miscellaneous', 'darkmod', fallback=True)
                app_settings['demo_mode'] = config.getboolean('Miscellaneous', 'DemoMode', fallback=False)
                app_settings['cold_storage_available'] = config.getboolean('Miscellaneous', 'Cold Storage Available', fallback=False)

    except Exception as e:
        print(f"Error loading configuration: {e}")
        print("Using default settings...")

    return app_settings


@debug_log
def update_gui(root, labels):
    global previous_prices, selected_coin, balances, is_tracker_active, after_id
    global ath_price_eu, ath_price_usd, ath_price_str, ath_coin_symbol, ath_cache

    #print("update_gui called")
    #print("is_tracker_active: ", is_tracker_active)

    # Stop if event is set, tracker is inactive, or root does not exist anymore
    # Should be inactive or False by Warm Storage, Cold Storage and Total Assets
    if stop_event.is_set() or not is_tracker_active or not root.winfo_exists():
        print("update_gui: stopped due to event, inactive state, or missing root window")

        return

    crypto = selected_coin.get() if selected_coin else None
    if not crypto:
        print("update_gui: no crypto selected, scheduling new update")
        # Cancel any existing scheduled update before setting a new one
        if after_id:
            root.after_cancel(after_id)
        after_id = root.after(10000, update_gui, root, labels)
        return

    selected_coin_str = labels['coins_dropdown'].get() # Get the string value
    #print(f"Selected coin in update_gui: {selected_coin_str}")
    ath_data = get_ath(selected_coin_str) # First call to get_ath

    if isinstance(ath_data, tuple) and len(ath_data) == 2:
        ath_price_usd = ath_data[0]
        ath_price_eur = ath_data[1]
     # Debug
    else:
        print(f"Error: Unexpected format for ATH data: {ath_data}")
        ath_price_eur = None
        ath_price_usd = None


    # Fetch the price ticker, EUR/USD rate, 24h change and balances

    ticker_data = get_crypto_ticker(crypto)
    rel_change = get_coin_change(crypto)

    eur_usd_rate = scrape_eur_usd()
    balances = get_warm_exchange_balance()
    # Removed the second call to get_ath here: ath_price_usd, ath_price_eur = get_ath(crypto)


    # Handle EUR rate as 1 if the selected coin is EUR. Of course... only for the Europeans.
    if crypto == 'EUR':
        eur_price = 1
        print('euro is 1')
    elif ticker_data:
        eur_price = ticker_data.get('eur_rate')

    else:
        eur_price = None

    # Ensure the data is valid
    if ticker_data and eur_usd_rate and balances:
        bal = balances.get(crypto, {})
        total_amount = bal.get('available', 0) + bal.get('in_order', 0)
        usd_price = eur_price * eur_usd_rate if eur_price is not None else None
        updated_time = time.strftime('%d-%m-%Y %H:%M:%S', time.localtime(ticker_data['updated']))
        if rel_change > 0:
            eur_arrow, eur_color = "▲", "green"
            usd_arrow, usd_color = "▲", "green"
        elif rel_change < 0:
            eur_arrow, eur_color = "▼", "red"
            usd_arrow, usd_color = "▼", "red"
        else:
            eur_arrow, eur_color = "", fg_color
            usd_arrow, usd_color = "", fg_color

        # Arrows for price direction (up or down)
        #eur_arrow, eur_color, usd_arrow, usd_color = "", fg_color, "", fg_color
        #if crypto in previous_prices and previous_prices[crypto]:
        #    pe, pu = previous_prices[crypto]['eur'], previous_prices[crypto]['usd']
        #    eur_arrow, eur_color = ("▲", "green") if eur_price > pe else ("▼", "red") if eur_price < pe else ("", fg_color)
        #    usd_arrow, usd_color = ("▲", "green") if usd_price > pu else ("▼", "red") if usd_price < pu else ("", fg_color)


        # Save the current prices as previous for the next comparison
        #previous_prices[crypto] = {'eur': eur_price, 'usd': usd_price}

        # Update GUI elements safely (only if they still exist)

        icon=get_coin_icon_main(crypto)
        #print(crypto)

        if 'header_white' in labels and tk.Frame.winfo_exists(labels['header_white'].master):
            labels['header_white'].config(text="Current", font=("Helvetica", 22, "bold"))


        if 'header_orange' in labels and tk.Frame.winfo_exists(labels['header_orange'].master):
            labels['header_orange'].config(text=f"{crypto} ", font=("Helvetica", 22, "bold"))

            # Get the ICON and set the label
            if icon:  # Make sure the ICON is not none
                labels['header_orange'].config(image=icon, compound="right")
                labels['header_orange'].image = icon  # Voorkom dat de afbeelding verdwijnt door garbage collection

        #if 'header_orange' in labels and tk.Frame.winfo_exists(labels['header_orange'].master):
            #labels['header_orange'].config(text=f"{crypto} ({coin_symbols.get(crypto, '')})", font=("Helvetica", 22, "bold"))
        #    labels['header_orange'].config(text=f"{crypto} {icon}", font=("Helvetica", 22, "bold"))
        if 'eur_text' in labels and tk.Frame.winfo_exists(labels['eur_text'].master):
            labels['eur_text'].config(text="EUR:", font=("Helvetica", 16))
        if 'eur_value' in labels and tk.Frame.winfo_exists(labels['eur_value'].master):

            eur_text = f"€{eur_price:.2f}" if eur_price is not None else "Failed"




            if eur_price is not None and eur_price < 100000:
                eur_text = f"€  {eur_price:.2f}"
                # Added extra space here
            labels['eur_value'].config(
                text=f"{eur_text} {eur_arrow}",
                fg=eur_color, font=("Helvetica", 16))
        if 'usd_text' in labels and tk.Frame.winfo_exists(labels['usd_text'].master):
            labels['usd_text'].config(text="USD:", font=("Helvetica", 16))
        if 'usd_value' in labels and tk.Frame.winfo_exists(labels['usd_value'].master):
            usd_text = f"${usd_price:.2f}" if usd_price is not None else "Failed"
            if usd_price is not None and usd_price < 100000:
                usd_text = f"$  {usd_price:.2f}"
                # Added extra space here
            labels['usd_value'].config(text=f"{usd_text} {usd_arrow}" if usd_price is not None else "Failed", fg=usd_color, font=("Helvetica", 16))
        if 'footer_text' in labels and tk.Frame.winfo_exists(labels['footer_text'].master):
            labels['footer_text'].config(text="Updated:", font=("Helvetica", 16))
        if 'footer_date' in labels and tk.Frame.winfo_exists(labels['footer_date'].master):
            labels['footer_date'].config(text=updated_time, fg="yellow", font=("Helvetica", 12))
        # Use the ath_data fetched at the beginning of the function
        if 'ath_label' in labels and tk.Widget.winfo_exists(labels['ath_label']) and 'ath_label_text' in labels:
            ath_label_var = labels['ath_label_text']
            usd_display = f"${ath_price_usd:.2f}" if ath_price_usd is not None else "N/A"
            eur_display = f"€{ath_price_eur:.2f}" if ath_price_eur is not None else "N/A"
            ath_label_var.set(f"Ath: {eur_display} / {usd_display}")
            #print(f"ATH Label StringVar set to: {ath_label_var.get()}")
            # more persitent way
            ath_cache = {
            "eur": ath_price_eur,
            "usd": ath_price_usd} if ath_price_eur is not None and ath_price_usd is not None else {"eur": 0.0, "usd": 0.0}


    else:
        # If fetching data failed, update labels accordingly
        if 'eur_value' in labels and tk.Frame.winfo_exists(labels['eur_value'].master):
            labels['eur_value'].config(text="Failed to retrieve data.", fg="red")
        if 'usd_value' in labels and tk.Frame.winfo_exists(labels['usd_value'].master):
            labels['usd_value'].config(text="Failed to retrieve data.", fg="red")
        if 'footer_date' in labels and tk.Frame.winfo_exists(labels['footer_date'].master):
            labels['footer_date'].config(text="Failed", fg="red")

    # Cancel the existing scheduled update before setting the new one
    if after_id:
        root.after_cancel(after_id)

    # Schedule the next update after 15 seconds
    after_id = root.after(par_refresh_main, update_gui, root, labels) # update every 30 seconds






@debug_log
def get_cold_storage_balance():
    global par_demo_mode
    cold_storage = {}
    print(par_demo_mode)
    if par_demo_mode:
        # Coins and amounts for demo_mode
        cold_storage = {
            'BTC': 0.1,
            'ETH': 10,
            'XRP': 10,
            'ADA': 10,
            'SOL': 10,
            'POLS': 10
        }
        return cold_storage  # Sla Excel inlezen over

    try:
        wb = openpyxl.load_workbook('tracker.xlsx')
        if 'Cold_Storage' in wb.sheetnames:
            Cold_Storage_ws = wb['Cold_Storage']
            row_num = 3
            while True:
                coin = Cold_Storage_ws.cell(row=row_num, column=1).value
                amount_str = Cold_Storage_ws.cell(row=row_num, column=2).value

                if not coin:
                    break

                try:
                    amount = float(amount_str)
                    if amount > 0:
                        cold_storage[coin] = amount
                except (ValueError, TypeError):
                    print(f"Warning: Invalid amount in Cold_Storage on row {row_num}, col B: {amount_str}. This loine will be skipped.")

                row_num += 1
        else:
            print("Warning: Sheet 'Cold_Storage' not found in tracker.xlsx. No cold storage data loaded.")
    except FileNotFoundError:
        print("Error: 'tracker.xlsx' not found. No cold storage data loaded.")
    except Exception as e:
        print(f"Error reading 'Cold_Storage' sheet: {e}. No cold storage data loaded.")

    return cold_storage

def init_excel():
    print("Init Excel Selected")
    # highlight_menu("Config", "Init Excel") # highlight_menu is not defined
    pass


def add_stocks():
    print("Add Stocks Selected")
    # highlight_menu("Config", "Add Stocks") # highlight_menu is not defined
    pass


def about():
    print("About")
    # highlight_menu("About", "About") # highlight_menu is not defined
    pass


def show_combined_storage(root, main_widgets):
    global is_tracker_active, updater_job_warm, status_label_warm, btc_label, back_button
    root.title("Combined Storage - Crypto Price Tracker V1.5")
    icon_path = os.path.join(os.getcwd(), "crypto", f"cws.ico")
    root.iconbitmap(icon_path)  # Your .ico file path here
    root.configure(bg=bg_color)

    is_tracker_active = False
    updater_job_warm = None
    status_label_warm = None
    btc_label = None
    back_button = None

    for menu in menubar.children.values():  # Iterate through all menus

        for i in range(menu.index('end') + 1):  # Loop through each item
            menu.entryconfig(i, state="disabled")  # Disablecall each item

    def update_combined_storage():
        global updater_job_warm, status_label_warm, btc_label, back_button, fg_ani, current_warm_data

        warm_balances = get_warm_exchange_balance()
        cold_balances = get_cold_storage_balance()

        combined = {}

        # Warm balances: available + in_order
        for coin, values in warm_balances.items():
            total = values.get('available', 0) + values.get('in_order', 0)
            combined[coin] = total

        # Cold balances: just add
        for coin, amount in cold_balances.items():
            combined[coin] = combined.get(coin, 0) + amount

        # Get prices
        prices = {}
        for coin in combined:
            try:
                if coin == "EUR":
                    prices[coin] = {"eur_rate": 1.0}
                else:
                    ticker = get_crypto_ticker(coin)
                    prices[coin] = ticker if ticker and 'eur_rate' in ticker else {"eur_rate": None}
            except Exception as e:
                logging.error(f"Error fetching price for {coin}: {e}")
                prices[coin] = {"eur_rate": None}

    # Build result
        result = []
        for coin, amount in combined.items():
            rate = prices.get(coin, {}).get("eur_rate", None)
            total_eur = round(amount * rate, 2) if rate is not None else None
            result.append({
                'Coin': coin,
                'Amount': round(amount, 8),
                'Rate': round(rate, 2),
                'Value': round(total_eur,2)

        })


        # Sorted Alfabetically
        displayed_coins = sorted(result, key=lambda x: x['Coin'])



        for widget in root.winfo_children():
            if widget not in [status_label_warm, btc_label, back_button] and not isinstance(widget, tk.Menu):
                widget.destroy()

        root.geometry("700x700")
        root.configure(bg=bg_color)

        # Title
        if par_demo_mode:
            both_storage_frame = tk.LabelFrame(
                root,
                text="Crypto Storage Assets ** Demo Mode **",
                font=("Helvetica", 18, "bold"),
                fg=fg_color,
                bg=bg_color,
                bd=2,
                relief="groove",
                labelanchor="n"
    )
        else:
            both_storage_frame = tk.LabelFrame(
                root,
                text="Crypto Storage Assets",
                font=("Helvetica", 18, "bold"),
                fg=fg_color,
                bg=bg_color,
                bd=2,
                relief="groove",
                labelanchor="n"
    )

        both_storage_frame.pack(pady=10, fill="x", padx=20)
        grid_container_frame = tk.Frame(both_storage_frame, bg=bg_color)
        grid_container_frame.pack(fill="x", padx=10, pady=5)

        # Configure columns for uniform sizing and alignment
        # Column 0: Icon (left aligned)
        grid_container_frame.grid_columnconfigure(0, weight=0, uniform="both_cols") # Don't let it expand too much
        # Column 1: Coin (left aligned)
        grid_container_frame.grid_columnconfigure(1, weight=1, uniform="both_cols")
        # Column 2: Rate (right aligned)
        grid_container_frame.grid_columnconfigure(2, weight=1, uniform="both_cols", minsize=100)
        # Column 3: Amount (right aligned)
        grid_container_frame.grid_columnconfigure(3, weight=1, uniform="both_cols", minsize=100)
        # Column 4: Value (right aligned)
        grid_container_frame.grid_columnconfigure(4, weight=1, uniform="both_cols", minsize = 100)
        grid_container_frame.grid_columnconfigure(5, weight=1, uniform="both_cols", minsize = 100)
        # Column 4: Value (right aligned)
        grid_container_frame.grid_columnconfigure(6, weight=1, uniform="both_cols")



        # Header row
        tk.Label(grid_container_frame, text="", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=0, sticky="w", padx=(0,2))
        tk.Label(grid_container_frame, text="Coin", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=1, sticky="w", padx=(0,2))
        tk.Label(grid_container_frame, text="Rate (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=2, sticky="e", padx=(10,10))
        tk.Label(grid_container_frame, text="Amount", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=3, sticky="e", padx=(10,10))
        tk.Label(grid_container_frame, text="Value (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=4, sticky="e", padx=(10,10))
        tk.Label(grid_container_frame, text="%▼▲", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=5, sticky="e", padx=(0,2))
        tk.Label(grid_container_frame, text="24H", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=6, sticky="e", padx=(0,2))

        total_eur_value=0
        # Data rows
        row_num = 1 # Start data rows from row 1 (after header)
        separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
        separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
        row_num += 1

        for entry in displayed_coins:
            both_coin = entry['Coin']
            both_amount = entry['Amount']
            both_rate = entry['Rate']
            both_value = entry['Value']




            if both_value > 1:
                #total_amount = balance_data['available'] + balance_data['in_order']
                total_eur_value=total_eur_value+both_value
                print(f"Check market: {both_coin}")
                both_change=get_coin_change(both_coin)

                if both_change > 0:
                    coin_change = "▲"
                    fg_change_color = "green"
                elif both_change < 0:
                    coin_change = "▼"
                    fg_change_color = "red"
                else:
                    coin_change =""

                icon = get_coin_icon(both_coin)
                symbol_label = tk.Label(grid_container_frame, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color)


                if icon:  # Check if the Icon has been loaded
                    symbol_label.config(image=icon, compound="left")
                    symbol_label.image = icon # Keep a reference!
                    symbol_label.grid(row=row_num, column=0, sticky="w", padx=(0,0))

                    tk.Label(grid_container_frame, text=both_coin, font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=1, sticky="w", padx=(0,2))
                    tk.Label(grid_container_frame, text=f"{both_rate:.2f}" , font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=2, sticky="e", padx=(10,10))
                    tk.Label(grid_container_frame, text=f"{both_amount:.2f}", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=3, sticky="e", padx=(10,10))
                    tk.Label(grid_container_frame, text=f"{both_value:.2f}" if both_value else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=4, sticky="e", padx=(10,2))
                    tk.Label(grid_container_frame, text=f"{both_change:.2f}" if both_change else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=5, sticky="e", padx=(10,2))
                    tk.Label(grid_container_frame, text=f"{coin_change}" if both_value else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=6, sticky="e", padx=(0,2))



                row_num += 1


        # Separator line
        # Place the separator in the grid_container_frame as well, spanning across columns
        separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
        separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
        row_num += 1
        tk.Label(
            grid_container_frame,
            text=f"Total Crypto Storage Value: €{total_eur_value:.2f}",
            font=("Helvetica", 12, "bold"),
            fg=fg_tot_storage,
            bg=bg_color
            ).grid(row=row_num, column=0, columnspan=7, sticky="e", pady=(5,0))



        if btc_label is None or not btc_label.winfo_exists():
            btc_label = tk.Label(root, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color, anchor="sw")
            btc_label.place(x=195, y=660)

        if back_button is None or not back_button.winfo_exists():
            icon_path = os.path.join(os.getcwd(), "crypto", f"back_blue.png")
            img = Image.open(icon_path).resize((20, 20))
            img_tk = ImageTk.PhotoImage(img)

            back_button = Button(
                root,
                image=img_tk,
                command=back_to_main_warm,  # Pass the function reference, not its call
                bg="grey",
                activebackground="forestgreen",
                highlightbackground="white",
                # Remove bd=0 or set it to a value > 0 to see relief
                bd=5,  # Set border width to 5 for visibility
                relief="raised",# Now relief will be visible
                width = 35
            )
            back_button.bind("<Enter>", on_hover)
            back_button.bind("<Leave>", on_leave)
            back_button.place(x=630, y=660)
            back_button.image = img_tk

        animate_status()
        updater_job_warm = root.after(par_refresh_warm, update_combined_storage)

    def animate_status():

        symbols = ["🔄", "🔃"]
        frame_interval = 300
        total_animation_time = 3000

        def animate(frame_idx=0, elapsed=0):
            if elapsed < total_animation_time:
                if status_label_warm and status_label_warm.winfo_exists():
                    status_label_warm.config(text=symbols[frame_idx % len(symbols)], fg=fg_cyan)
                    if btc_label and btc_label.winfo_exists():
                        btc_label.config(text="")
                root.after(frame_interval, animate, frame_idx + 1, elapsed + frame_interval)
            else:
                if status_label_warm and status_label_warm.winfo_exists():
                    status_label_warm.config(text="✅", fg=fg_ani)
                try:
                    btc_val = get_coin_exchange_ticker('BTC-EUR')
                    btc_price = btc_val.get("price", "N/A")
                    btc_price_str = btc_val.get("price", "N/A") # Get it as a string

                    display_btc_price = "N/A" # Default to "N/A"

                    if isinstance(btc_price_str, str) and btc_price_str != "N/A":
                        try:
                            btc_price_float = float(btc_price_str)
                            formatted_btc_price = f"{btc_price_float:.2f}"
                        except ValueError:
                            # Handle cases where the string isn't a valid float
                            # For example, if it's "Error" or an empty string
                            formatted_btc_price = "Error" # Or any other appropriate message

                    if btc_label and btc_label.winfo_exists():
                            eur_usd_rate = scrape_eur_usd()

                            try:
                                btc_price_usd = float(btc_price) * eur_usd_rate
                                formatted_price = str(round(btc_price_usd))

                            except ValueError:
                                print("Error: btc_price is not a valid number!")

                            btc_label.config(text=" Current Bitcoin Price: € " + btc_price + " / $ " + formatted_price, fg=fg_color)



                except Exception as e:
                    logging.error(f"BTC price fetch failed: {e}")

        animate()

    def back_to_main_warm():
        global is_tracker_active, updater_job_warm, current_warm_data
        #print("Status of write_warm", par_write_warm)
        # Write current warm storage data to spreadsheet before going back
        try:
            if current_warm_data and par_write_warm:
                write_horizontal(current_warm_data,"warm")
                print(f"Exported {len(current_warm_data)} coins to spreadsheet")
            else:
                print("No warm storage data to export")
        except Exception as e:
            print(f"Error writing to spreadsheet: {e}")
            logging.error(f"Spreadsheet export failed: {e}")


        root.title("Main - Crypto Price Tracker V1.5")
        icon_path = os.path.join(os.getcwd(), "crypto", f"MoB.ico")
        root.iconbitmap(icon_path)  # Your .ico file path here
        root.configure(bg=bg_color)
        global is_tracker_active, updater_job_warm
        is_tracker_active = True
        for menu in menubar.children.values():
            for i in range(menu.index('end') + 1):
                menu.entryconfig(i, state="normal")

        if updater_job_warm:
            root.after_cancel(updater_job_warm)
            updater_job_warm = None
        for widget in root.winfo_children():
            if not isinstance(widget, tk.Menu):
                widget.destroy()
        show_main_screen(root)

    # Clear current widgets
    for widget in root.winfo_children():
        if not isinstance(widget, tk.Menu):
            widget.pack_forget()
            widget.place_forget()
            widget.destroy()

    # Status label
    status_label_warm = tk.Label(root, text="", font=("Helvetica", 18), fg="orange", bg=bg_color, anchor="sw")
    status_label_warm.place(x=20, y=660)





    update_combined_storage()





@debug_log
def show_warm_storage(root):
    global is_tracker_active, updater_job_warm, status_label_warm, btc_label, back_button
    root.title("Warm Storage - Crypto Price Tracker V1.5")
    icon_path = os.path.join(os.getcwd(), "crypto", f"ThermoWarm.ico")
    root.iconbitmap(icon_path)  # Your .ico file path here
    root.configure(bg=bg_color)

    is_tracker_active = False
    updater_job_warm = None
    status_label_warm = None
    btc_label = None
    back_button = None

    for menu in menubar.children.values():  # Iterate through all menus

        for i in range(menu.index('end') + 1):  # Loop through each item
            menu.entryconfig(i, state="disabled")  # Disablecall each item

    def update_warm_storage():
        global updater_job_warm, status_label_warm, btc_label, back_button, fg_ani, current_warm_data

        balances = get_warm_exchange_balance()
        prices = {}

        for coin in balances:
            try:
                if coin == "EUR":
                    prices[coin] = {"eur_rate": 1.0}
                else:
                    ticker = get_crypto_ticker(coin)
                    prices[coin] = ticker if ticker and 'eur_rate' in ticker else {"eur_rate": None}
            except Exception as e:
                logging.error(f"Error fetching price for {coin}: {e}")
                prices[coin] = {"eur_rate": None}

        # Clean screen, preserve essential widgets
        for widget in root.winfo_children():
            if widget not in [status_label_warm, btc_label, back_button] and not isinstance(widget, tk.Menu):
                widget.destroy()

        root.geometry("700x700")
        root.configure(bg=bg_color)

        # Title
        if par_demo_mode:
            warm_storage_frame = tk.LabelFrame(
                root,
                text="Warm Storage Assets ** Demo Mode **",
                font=("Helvetica", 18, "bold"),
                fg="orange",
                bg=bg_color,
                height=600,
                bd=2,
                relief="groove",
                labelanchor="n"
    )
        else:
            warm_storage_frame = tk.LabelFrame(
                root,
                text="Warm Storage Assets",
                font=("Helvetica", 18, "bold"),
                fg="orange",
                bg=bg_color,
                height=600,
                bd=2,
                relief="groove",
                labelanchor="n"
    )

        #warm_storage_frame.pack(pady=10, fill="x", padx=20)
        warm_storage_frame.pack(fill="both", expand=False, padx=20, pady=10)
        warm_storage_frame.pack_propagate(False) # makes sure the height will be kept

        if balances:
            sorted_balances = sorted(balances.items())
            displayed_coins = []

            #symbol_width = 10
            #coin_width = len("Coin")+2
            #price_width = len("Rate (EUR)")
            #amount_width = len("Amount Coins")
            #value_width = len("Value (EUR)")

            # Create a frame to hold the grid-based content
            grid_container_frame = tk.Frame(warm_storage_frame, bg=bg_color)
            grid_container_frame.pack(fill="x", padx=10, pady=5)

            # Configure columns for uniform sizing and alignment
            # Column 0: Icon (left aligned)
            grid_container_frame.grid_columnconfigure(0, weight=0, uniform="warm_cols") # Don't let it expand too much
            # Column 1: Coin (left aligned)
            grid_container_frame.grid_columnconfigure(1, weight=1, uniform="warm_cols")
            # Column 2: Rate (right aligned)
            grid_container_frame.grid_columnconfigure(2, weight=1, uniform="warm_cols", minsize=100)
            # Column 3: Amount (right aligned)
            grid_container_frame.grid_columnconfigure(3, weight=1, uniform="warm_cols", minsize=100)
            # Column 4: Value (right aligned)
            grid_container_frame.grid_columnconfigure(4, weight=1, uniform="warm_cols", minsize = 100)
            # Column 4: Value (right aligned)
            grid_container_frame.grid_columnconfigure(5, weight=1, uniform="warm_cols", minsize = 100)
            grid_container_frame.grid_columnconfigure(6, weight=1, uniform="warm_cols")


            # Header row
            tk.Label(grid_container_frame, text="", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=0, sticky="w", padx=(0,2))
            tk.Label(grid_container_frame, text="Coin", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=1, sticky="w", padx=(0,2))
            tk.Label(grid_container_frame, text="Rate (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=2, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="Amount", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=3, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="Value (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=4, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="%▼▲", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=5, sticky="e", padx=(0,2))
            tk.Label(grid_container_frame, text="24H", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=6, sticky="e", padx=(0,2))


            row_num = 1 # Start line from row 2 (after header)
            separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
            separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
            # Clear previous data and prepare new data for spreadsheet
            # Data rows
            row_num += 1 # Start data rows from row 2 (after header)
            current_warm_data = []

            for coin, balance_data in sorted_balances:
                available = balance_data['available']
                in_order = balance_data['in_order']
                total_amount = available + in_order
                eur_price = prices.get(coin, {}).get('eur_rate')
                eur_value = total_amount * eur_price if eur_price is not None else None

                if eur_value is not None and eur_value >= 0.1:
                    displayed_coins.append((coin, balance_data, eur_price, eur_value))

                    # Store data for spreadsheet export
                    current_warm_data.append({
                        'coin': coin,
                        'amount': total_amount,
                        'rate': eur_price,
                        'value': eur_value
                    })

            # Coin rows
            for coin, balance_data, eur_price, eur_value in displayed_coins:
                total_amount = balance_data['available'] + balance_data['in_order']
                icon = get_coin_icon(coin)
                symbol_label = tk.Label(grid_container_frame, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color)
                #row_frame = tk.Frame(root, bg=bg_color)
                warm_change=get_coin_change(coin)

                if warm_change > 0:
                    coin_change = "▲"
                    fg_change_color = "green"
                elif warm_change < 0:
                    coin_change = "▼"
                    fg_change_color = "red"
                else:
                    coin_change = "-"
                    fg_change_color = fg_color



                # Label with coin-name + icon
                #coin_label = tk.Label(row_frame, text=coin, font=("Helvetica", 12), fg=fg_color, bg=bg_color, width=coin_width, anchor="w")

                #ICON_COIN_WARM

                if icon:  # Check if the Icon has been loaded
                    symbol_label.config(image=icon, compound="left")
                    symbol_label.image = icon # Keep a reference!
                symbol_label.grid(row=row_num, column=0, sticky="w", padx=(0,2))

                tk.Label(grid_container_frame, text=coin, font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=1, sticky="w", padx=(0,2))
                tk.Label(grid_container_frame, text=f"{eur_price:.2f}" if eur_price else "N/A", font=("Helvetica", 12), fg=fg_change_color if eur_price else "red", bg=bg_color).grid(row=row_num, column=2, sticky="e", padx=(10,10))
                tk.Label(grid_container_frame, text=f"{total_amount:.2f}", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=3, sticky="e", padx=(10,10))
                tk.Label(grid_container_frame, text=f"€{eur_value:.2f}" if eur_value else "N/A", font=("Helvetica", 12), fg=fg_change_color if eur_value else "red", bg=bg_color).grid(row=row_num, column=4, sticky="e", padx=(10,2))
                tk.Label(grid_container_frame, text=f"{warm_change:.2f}" if eur_value else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=5, sticky="e", padx=(0,2))
                tk.Label(grid_container_frame, text=f"{coin_change}" if eur_value else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=6, sticky="e", padx=(0,2))

                row_num += 1

            # Separator line
            # Place the separator in the grid_container_frame as well, spanning across columns
            separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
            separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
            row_num += 1



            # Totals
            total_eur_value = sum((item[1]['available'] + item[1]['in_order']) * item[2]
                                 for item in displayed_coins if item[2] is not None)
            sep0_frame = tk.Frame(root, bg=bg_color)
            sep0_frame.pack(pady=5, fill="x") # This frame contains the line


            tk.Label(
                grid_container_frame,
                text=f"Total Warm Storage Value: €{total_eur_value:.2f}",
                font=("Helvetica", 12, "bold"),
                fg=fg_tot_storage,
                bg=bg_color
                ).grid(row=row_num, column=0, columnspan=7, sticky="e", pady=(5,0))




        else:
            tk.Label(root, text="No Assets Found.", font=("Helvetica", 16), fg=fg_color, bg=bg_color).pack()

        # Persistent widgets
        if btc_label is None or not btc_label.winfo_exists():
            btc_label = tk.Label(root, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color, anchor="sw")
            btc_label.place(x=195, y=660)

        if back_button is None or not back_button.winfo_exists():
            icon_path = os.path.join(os.getcwd(), "crypto", f"back_blue.png")
            img = Image.open(icon_path).resize((20, 20))
            img_tk = ImageTk.PhotoImage(img)

            back_button = Button(
                root,
                image=img_tk,
                command=back_to_main_warm,  # Pass the function reference, not its call
                bg="grey",
                activebackground="forestgreen",
                highlightbackground="white",
                # Remove bd=0 or set it to a value > 0 to see relief
                bd=5,  # Set border width to 5 for visibility
                relief="raised",# Now relief will be visible
                width = 35
            )
            back_button.bind("<Enter>", on_hover)
            back_button.bind("<Leave>", on_leave)
            back_button.place(x=630, y=660)
            back_button.image = img_tk

        animate_status()
        updater_job_warm = root.after(par_refresh_warm, update_warm_storage)

    def animate_status():

        symbols = ["🔄", "🔃"]
        frame_interval = 300
        total_animation_time = 3000

        def animate(frame_idx=0, elapsed=0):
            if elapsed < total_animation_time:
                if status_label_warm and status_label_warm.winfo_exists():
                    status_label_warm.config(text=symbols[frame_idx % len(symbols)], fg=fg_cyan)
                    if btc_label and btc_label.winfo_exists():
                        btc_label.config(text="")
                root.after(frame_interval, animate, frame_idx + 1, elapsed + frame_interval)
            else:
                if status_label_warm and status_label_warm.winfo_exists():
                    status_label_warm.config(text="✅", fg=fg_ani)
                try:
                    btc_val = get_coin_exchange_ticker('BTC-EUR')
                    btc_price = btc_val.get("price", "N/A")
                    btc_price_str = btc_val.get("price", "N/A") # Get it as a string

                    display_btc_price = "N/A" # Default to "N/A"

                    if isinstance(btc_price_str, str) and btc_price_str != "N/A":
                        try:
                            btc_price_float = float(btc_price_str)
                            formatted_btc_price = f"{btc_price_float:.2f}"
                        except ValueError:
                            # Handle cases where the string isn't a valid float
                            # For example, if it's "Error" or an empty string
                            formatted_btc_price = "Error" # Or any other appropriate message

                    if btc_label and btc_label.winfo_exists():
                            eur_usd_rate = scrape_eur_usd()

                            try:
                                btc_price_usd = float(btc_price) * eur_usd_rate
                                formatted_price = str(round(btc_price_usd))

                            except ValueError:
                                print("Error: btc_price is not a valid number!")

                            btc_label.config(text=" Current Bitcoin Price: € " + btc_price + " / $ " + formatted_price, fg=fg_color)



                except Exception as e:
                    logging.error(f"BTC price fetch failed: {e}")

        animate()

    def back_to_main_warm():
        global is_tracker_active, updater_job_warm, current_warm_data
        #print("Status of write_warm", par_write_warm)
        # Write current warm storage data to spreadsheet before going back
        try:
            if current_warm_data and par_write_warm:
                write_horizontal(current_warm_data,"warm")
                print(f"Exported {len(current_warm_data)} coins to spreadsheet")
            else:
                print("No warm storage data to export")
        except Exception as e:
            print(f"Error writing to spreadsheet: {e}")
            logging.error(f"Spreadsheet export failed: {e}")


        root.title("Main - Crypto Price Tracker V1.5")
        icon_path = os.path.join(os.getcwd(), "crypto", f"MoB.ico")
        root.iconbitmap(icon_path)  # Your .ico file path here
        root.configure(bg=bg_color)
        global is_tracker_active, updater_job_warm
        is_tracker_active = True
        for menu in menubar.children.values():
            for i in range(menu.index('end') + 1):
                menu.entryconfig(i, state="normal")

        if updater_job_warm:
            root.after_cancel(updater_job_warm)
            updater_job_warm = None
        for widget in root.winfo_children():
            if not isinstance(widget, tk.Menu):
                widget.destroy()
        show_main_screen(root)

    # Clear current widgets
    for widget in root.winfo_children():
        if not isinstance(widget, tk.Menu):
            widget.pack_forget()
            widget.place_forget()
            widget.destroy()

    # Status label
    status_label_warm = tk.Label(root, text="", font=("Helvetica", 18), fg="orange", bg=bg_color, anchor="sw")
    status_label_warm.place(x=20, y=660)

    update_warm_storage()

def restart_program():
    """Restarts the current program.
    Note: this function does not return. Any cleanup action (like
    saving data) must be done before calling this function.
    """
    python = sys.executable
    os.execv(python, [python] + sys.argv)

@debug_log
def show_cold_storage(root, main_widgets):
    # Globals to manage tracker status, update job, and status label
    global is_tracker_active, updater_job_cold, status_label_cold, btc_label, back_button

    root.title("Cold Storaqge - Crypto Price Tracker V1.5")
    icon_path = os.path.join(os.getcwd(), "crypto", f"ThermoCold.ico")
    print(icon_path)
    root.iconbitmap(icon_path)  # Your .ico file path here
    root.configure(bg=bg_color)
    is_tracker_active = False
    updater_job_cold = None
    status_label_cold = None
    btc_label = None
    back_button = None
    for menu in menubar.children.values():  # Iterate through all menus
        for i in range(menu.index('end') + 1):  # Loop through each item
            menu.entryconfig(i, state="disabled")  # Disable each item

    def update_cold_storage():
        """Refresh cold storage display and animate status"""
        global updater_job_cold, status_label_cold, btc_label, back_button, fg_ani
        global current_cold_data

        # Get updated cold storage balances and latest prices
        cold_storage_balances = get_cold_storage_balance()
        prices = {}
        for coin in cold_storage_balances:
            prices[coin] = get_crypto_ticker(coin)

        # Destroy all widgets except status_label_cold and Menu widgets
        for widget in root.winfo_children():
            if widget not in [status_label_cold] and not isinstance(widget, tk.Menu):
                widget.destroy()

        # Set window properties
        root.geometry("700x700")
        root.configure(bg=bg_color)

        current_cold_data = []

        # Display Cold Storage header
        if par_demo_mode:
            cold_storage_frame = tk.LabelFrame(
                root,
                text="Cold Storage Assets ** Demo Mode **",
                font=("Helvetica", 18, "bold"),
                fg=fg_cold,
                bg=bg_color,
                height=600,
                bd=2,
                relief="groove",
                labelanchor="n"
                )

        else:
            cold_storage_frame = tk.LabelFrame(
                root,
                text="Cold Storage Assets",
                font=("Helvetica", 18, "bold"),
                fg=fg_cold,
                bg=bg_color,
                height=600,
                bd=2,
                relief="groove",
                labelanchor="n"
                )

        cold_storage_frame.pack(fill="both", expand=False, padx=20, pady=10)
        cold_storage_frame.pack_propagate(False) # makes sure the height will be kept
        #cold_storage_frame.pack(pady=10, fill="x", padx=20)


        if cold_storage_balances:
            sorted_balances = sorted(cold_storage_balances.items())

            # Create a frame to hold the grid-based content
            grid_container_frame = tk.Frame(cold_storage_frame, bg=bg_color)
            grid_container_frame.pack(fill="x", padx=10, pady=5)

            # Configure columns for uniform sizing and alignment
            # Column 0: Icon (left aligned)
            grid_container_frame.grid_columnconfigure(0, weight=0, uniform="cold_cols") # Don't let it expand too much
            # Column 1: Coin (left aligned)
            grid_container_frame.grid_columnconfigure(1, weight=1, uniform="cold_cols")
            # Column 2: Rate (right aligned)
            grid_container_frame.grid_columnconfigure(2, weight=1, uniform="cold_cols", minsize=100)
            # Column 3: Amount (right aligned)
            grid_container_frame.grid_columnconfigure(3, weight=1, uniform="cold_cols", minsize=100)
            # Column 4: Value (right aligned)
            grid_container_frame.grid_columnconfigure(4, weight=1, uniform="cold_cols", minsize = 100)
            grid_container_frame.grid_columnconfigure(5, weight=1, uniform="cold_cols", minsize = 100)
            # Column 4: Value (right aligned)
            grid_container_frame.grid_columnconfigure(6, weight=1, uniform="cold_cols")



            # Header row

            tk.Label(grid_container_frame, text="", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=0, sticky="w", padx=(0,2))
            tk.Label(grid_container_frame, text="Coin", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=1, sticky="w", padx=(0,2))
            tk.Label(grid_container_frame, text="Rate (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=2, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="Amount", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=3, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="Value (€)", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=4, sticky="e", padx=(10,10))
            tk.Label(grid_container_frame, text="%▼▲", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=5, sticky="e", padx=(0,2))
            tk.Label(grid_container_frame, text="24H", font=("Helvetica", 14, "bold"), fg=fg_color, bg=bg_color).grid(row=0, column=6, sticky="e", padx=(0,2))
            row_num = 1 # Start line from row 2 (after header)
            separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
            separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
            # Data rows
            row_num += 1 # Start data rows from row 2 (after header)


            for coin, amount in sorted_balances:
                eur_price = prices.get(coin, {}).get('eur_rate')
                eur_value = amount * eur_price if eur_price is not None else None
                icon = get_coin_icon(coin)
                cold_change=get_coin_change(coin)

                if cold_change > 0:
                    coin_change = "▲"
                    fg_change_color = "green"
                elif cold_change < 0:
                    coin_change = "▼"
                    fg_change_color = "red"
                else:
                    coin_change =""




                symbol_label = tk.Label(grid_container_frame, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color)
                if icon:
                    symbol_label.config(image=icon, compound="left")
                    symbol_label.image = icon # Keep a reference!
                symbol_label.grid(row=row_num, column=0, sticky="w", padx=(0,2))

                tk.Label(grid_container_frame, text=coin, font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=1, sticky="w", padx=(0,2))
                tk.Label(grid_container_frame, text=f"{eur_price:.2f}" if eur_price else "N/A", font=("Helvetica", 12), fg=fg_change_color if eur_price else "red", bg=bg_color).grid(row=row_num, column=2, sticky="e", padx=(10,10))
                tk.Label(grid_container_frame, text=f"{amount:.2f}", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=3, sticky="e", padx=(10,10))
                tk.Label(grid_container_frame, text=f"€{eur_value:.2f}" if eur_value else "N/A", font=("Helvetica", 12), fg=fg_change_color if eur_value else "red", bg=bg_color).grid(row=row_num, column=4, sticky="e", padx=(10,2))
                tk.Label(grid_container_frame, text=f"{cold_change:.2f}", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=5, sticky="e", padx=(10,2))
                tk.Label(grid_container_frame, text=f"{coin_change}" if eur_value else "N/A", font=("Helvetica", 12), fg=fg_change_color, bg=bg_color).grid(row=row_num, column=6, sticky="e", padx=(0,2))

                row_num += 1

            # Separator line
            # Place the separator in the grid_container_frame as well, spanning across columns
            separator = tk.Frame(grid_container_frame, height=1, bg=fg_color)
            separator.grid(row=row_num, column=0, columnspan=7, sticky="ew", pady=5)
            row_num += 1

            # Total value
            total_cold_value = sum(
                amount * prices.get(coin, {}).get('eur_rate', 0)
                for coin, amount in cold_storage_balances.items()
                if prices.get(coin, {}).get('eur_rate') is not None
                )

            # Place the total value label directly in the grid_container_frame
            # Span it across columns and stick to the east
            tk.Label(
                grid_container_frame,
                text=f"Total Cold Storage Value: €{total_cold_value:.2f}",
                font=("Helvetica", 12, "bold"),
                fg=fg_cold,
                bg=bg_color
                ).grid(row=row_num, column=0, columnspan=7, sticky="e", pady=(5,0))

        else:
            # No cold storage assets found
            no_assets_label = tk.Label(root, text="No Cold Storage Assets Found.", font=("Helvetica", 16), fg=fg_cold, bg=bg_color)
            no_assets_label.pack()

        # Persistent widgets
        if btc_label is None or not btc_label.winfo_exists():
            btc_label = tk.Label(root, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color, anchor="sw")
            btc_label.place(x=195, y=660)

        if back_button is None or not back_button.winfo_exists():
            icon_path = os.path.join(os.getcwd(), "crypto", f"back_blue.png")
            img = Image.open(icon_path).resize((20, 20))
            img_tk = ImageTk.PhotoImage(img)

            back_button = Button(
                root,
                image=img_tk,
                command=back_to_main_cold,  # Pass the function reference, not its call
                bg="grey",
                activebackground="forestgreen",
                highlightbackground="white",
                # Remove bd=0 or set it to a value > 0 to see relief
                bd=5,  # Set border width to 5 for visibility
                relief="raised", # Now relief will be visible
                width=35
            )
            back_button.bind("<Enter>", on_hover)
            back_button.bind("<Leave>", on_leave)
            back_button.place(x=630, y=660)
            back_button.image = img_tk

        # Start animation for status symbol
        animate_status()

        # Schedule next update after 10 seconds
        updater_job_cold = root.after(par_refresh_cold, update_cold_storage)

    def animate_status():
        """Animate the status label showing update progress"""
        symbols = ["🔄", "🔃"]  # Rotating update symbols
        frame_interval = 300   # Time between symbol changes
        total_animation_time = 3000  # Total animation duration (3 sec)
        elapsed = 0

        def animate(frame_idx=0, elapsed=0):

            if elapsed < total_animation_time:
                if status_label_cold is not None and status_label_cold.winfo_exists():
                    status_label_cold.config(text=f"{symbols[frame_idx % len(symbols)]}", fg=fg_cyan)
                root.after(frame_interval, animate, frame_idx + 1, elapsed + frame_interval)
            else:
                if status_label_cold and status_label_cold.winfo_exists():
                    status_label_cold.config(text="✅", fg=fg_ani)
                try:
                    btc_val = get_coin_exchange_ticker('BTC-EUR')
                    btc_price = btc_val.get("price", "N/A")
                    if btc_label and btc_label.winfo_exists():
                        eur_usd_rate = scrape_eur_usd()

                        try:
                            btc_price_usd = float(btc_price) * eur_usd_rate
                            formatted_price = str(round(btc_price_usd))

                        except ValueError:
                            print("Error: btc_price is not a valid number!")

                        btc_label.config(text=" Current Bitcoin Price: € " + btc_price + " / $ " + formatted_price, fg=fg_color)

                except Exception as e:
                    logging.error(f"BTC price fetch failed: {e}")
        animate()


    def back_to_main_cold():
        global is_tracker_active, updater_job_warm, current_cold_data
        # print("Status of write_cold", par_write_cold)
        # Write current cold storage data to spreadsheet before going back
        try:
            if current_cold_data and par_write_cold:
                write_horizontal(current_cold_data,"cold")
                print(f"Exported {len(current_cold_data)} coins to spreadsheet")
            else:
                print("No cold storage data to export")
        except Exception as e:
            print(f"Error writing to spreadsheet: {e}")
            logging.error(f"Spreadsheet export failed: {e}")


        root.title("Main - Crypto Price Tracker V1.5")
        icon_path = os.path.join(os.getcwd(), "crypto", f"MoB.ico")

        root.iconbitmap(icon_path)  # Your .ico file path here
        root.configure(bg=bg_color)
        global is_tracker_active, updater_job_cold
        is_tracker_active = True
        for menu in menubar.children.values():
            for i in range(menu.index('end') + 1):
                menu.entryconfig(i, state="normal")
        if updater_job_cold:
            root.after_cancel(updater_job_cold)
            updater_job_cold = None
        for widget in root.winfo_children():
            if not isinstance(widget, tk.Menu):
                widget.destroy()
        show_main_screen(root)

    # Clear current widgets
    for widget in root.winfo_children():
        if not isinstance(widget, tk.Menu):
            widget.pack_forget()
            widget.place_forget()
            widget.destroy()

    # Status label
    status_label_cold = tk.Label(root, text="", font=("Helvetica", 18), fg=fg_cyan, bg=bg_color, anchor="sw")
    status_label_cold.place(x=20, y=660)

    update_cold_storage()


@debug_log
def set_total_stocks(parent):
    global par_demo_mode
    if par_demo_mode:
        return


    filename='tracker.xlsx'

    def save_and_close():
        global total_stocks
        try:
            total_stocks = float(entry.get())
            print(f"Stocks saved: €{total_stocks:.2f}")
            # --- Excel Writing Logic ---
            try:
                wb = openpyxl.load_workbook(filename)
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            try:
                ws = wb['Stocks']
            except KeyError:
                ws = wb.create_sheet('Stocks')

            latest_row = max((cell.row for row in ws.iter_rows() for cell in row if cell.value), default=3)
            next_row = latest_row + 1

            ws['A' + str(next_row)] = today
            ws['B' + str(next_row)] = total_stocks

            wb.save(filename)
            #print(f"Data written to Excel.")
            # --- End Excel Writing Logic ---
            top.destroy()
        except ValueError:
            total_stocks = 0.0
            print("Invalid input, defaulting to €0.00")
            top.destroy()

    top = tk.Toplevel(parent)
    top.title("Input Total Stocks Value")
    top.geometry("300x100")
    top.grab_set()

    label = tk.Label(top, text="Enter total Stocks value in EUR:")
    label.pack(pady=5)

    entry = tk.Entry(top)
    entry.pack(pady=5)

    save_button = tk.Button(top, text="Save", command=save_and_close)
    save_button.pack(pady=5)

    parent.wait_window(top)




def get_coin_data(file_path, coin_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "CSV_History" not in wb.sheetnames:
        print("Worksheet 'CSV_History' not found.")
        return

    sheet = wb["CSV_History"]

    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[headers.get("Coin")] == coin_name:
            avg_buy = row[headers.get("Avg Buy €")]
            invested = row[headers.get("Invested €")]
            print(f"{coin_name} → Avg Buy €: {avg_buy}, Invested €: {invested}")
            return

    print(f"Coin '{coin_name}' not found.")




def find_eur_and_get_amounts(file_path):
    """
    Opens an Excel file, finds the row containing 'EUR' in the
    'CSV_History' worksheet, and returns the values from the
    'Invest' and 'Amount withdrawal' columns in that row.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        tuple: (success: bool, deposit_value, withdrawal_value)

    """
    global T_EUR_I
    global T_EUR_O
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['CSV_History']
      # Fixed position

        T_EUR_I = sheet["L2"].value
        T_EUR_O = sheet["M2"].value
        print(f"✅ Values Found:\n  In: {T_EUR_I}\n  Out: {T_EUR_O}")

        return True, T_EUR_I, T_EUR_O


    except FileNotFoundError:
        print(f"Error: File not found at '{file_name}'")
    except KeyError:
        print(f"Error: Worksheet 'CSV_History' not found in the file.")
    except Exception as e:
        print(f"An error occurred: {e}")
    return False, None, None


def on_hover(event):
    event.widget.config(bg="lightblue")

def on_leave(event):
    event.widget.config(bg="grey")

# Writes values Total Assets screen to tracker.xls sheet=
@debug_log
def write_totals(total_warm_value, total_cold_value, total_stocks_value, total_assets_value,
                 amount_deposit, amount_withdraw, t_invest, T_PL, pl_percentage, btc_price):
    global par_demo_mode

    if par_demo_mode:
        return

    file_name = "tracker.xlsx"
    """Opens or creates an Excel file and writes totals on the latest row + 1"""
    columns = ["Date", "Warm Storage", "Cold Storage", "Value Stocks", "Total Assets",
               "--------", "EURO In", "EURO Out", "Investment", "Total P/L", "Percentage",
               "----", "BTC EUR", "BTC USD"]

    try:
        # Try to open the workbook, create if it doesn't exist
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Check if worksheet exists, else create it
    if "Assets_History" in wb.sheetnames:
        ws = wb["Assets_History"]
    else:
        ws = wb.create_sheet("Assets_History")

        # Add column headers
        for col_index, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_index, value=col_name)

        # Set column width to 15
        for col_index in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_index)].width = 15

    # Find the latest filled row
    latest_row = max((cell.row for row in ws.iter_rows() for cell in row if cell.value), default=1)
    eur_usd_rate = scrape_eur_usd()
    def clean_numeric(value):
        """Removes currency symbols and converts to float"""
        if isinstance(value, str):
            return float(value.replace("€", "").replace(",", ""))
        return round(float(value),2)


    # Write data in the next available row
    next_row = latest_row + 1
    formatted_price_eur = clean_numeric(btc_price)
    formatted_price_usd = round((formatted_price_eur * eur_usd_rate),0)
    # Write values to the worksheet
    ws.cell(row=next_row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))  # Date & Time
    ws.cell(row=next_row, column=2, value=clean_numeric(total_warm_value))  # Warm Storage
    ws.cell(row=next_row, column=3, value=clean_numeric(total_cold_value)) # Cold Storage
    ws.cell(row=next_row, column=4, value=total_stocks_value)  # Value Stocks
    ws.cell(row=next_row, column=5, value=clean_numeric(total_assets_value))  # Total Assets
    ws.cell(row=next_row, column=7, value=clean_numeric(amount_deposit))  # EURO In
    ws.cell(row=next_row, column=8, value=clean_numeric(amount_withdraw))  # EURO Out
    ws.cell(row=next_row, column=9, value=t_invest)  # Investment
    ws.cell(row=next_row, column=10, value=clean_numeric(T_PL))  # Investment
    ws.cell(row=next_row, column=11, value=clean_numeric(pl_percentage))  # Percentage
    ws.cell(row=next_row, column=13, value=formatted_price_eur)  # Bitcoin Price
    ws.cell(row=next_row, column=14, value=formatted_price_usd) # Bitcoin Price






    # Save the workbook
    wb.save(file_name)

    #print(f"Data written to '{file_name}' in worksheet 'Assets_History' at row {next_row}")

# Writes values Warm Assets screen to tracker.xls sheet


@debug_log
def write_horizontal(coins_data, storage):
    global par_demo_mode
    """
    Writes warm or cold storage data horizontally across columns

    Args:
        coins_data: List of dictionaries with keys: 'coin', 'amount', 'rate', 'value'
                   Example: [{'coin': 'BTC', 'amount': 1.5, 'rate': 45000, 'value': 67500}, ...]
    """
    if par_demo_mode:
        return
    file_name = "tracker.xlsx"

    try:
        # Try to open the workbook, create if it doesn't exist
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # print (storage)
    if storage == "warm" :
        	# Check if worksheet exists, else create it
        if "Warm_History" in wb.sheetnames:
                ws = wb["Warm_History"]
        else:
                ws = wb.create_sheet("Warm_History")

    if storage == "cold" :
        # Check if worksheet exists, else create it
        if "Cold_History" in wb.sheetnames:
                ws = wb["Cold_History"]
        else:
                ws = wb.create_sheet("Cold_History")

        # Remove default sheet if it exists and is empty
    if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Find the next available row (first completely empty row)
    next_row = 1
    while any(ws.cell(row=next_row, column=col).value for col in range(1, 20)):  # Check first 20 columns
        next_row += 1

    def clean_numeric(value):
        """Removes currency symbols and converts to float"""
        if isinstance(value, str):
            return float(value.replace("€", "").replace(",", "").replace("$", ""))
        return float(value)

    # Write the layout as specified
    # First row: DATE in column 1, COINS in column 3
    ws.cell(row=next_row, column=1, value="DATE")
    ws.cell(row=next_row, column=3, value="COINS")

    # Second row: Current date in column 1, then coins starting from column 3
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=next_row + 1, column=1, value=current_date)

    # Write coin names starting from column 3
    for i, coin_data in enumerate(coins_data):
        ws.cell(row=next_row + 1, column=3 + i, value=coin_data['coin'])

    # Third row: "Rate" in column 2
    ws.cell(row=next_row + 2, column=2, value="Rate")

    # Write rates for each coin
    for i, coin_data in enumerate(coins_data):
        ws.cell(row=next_row + 2, column=3 + i, value=clean_numeric(coin_data['rate']))

    # Fourth row: "Amount" in column 2
    ws.cell(row=next_row + 3, column=2, value="Amount")

    # Write amounts for each coin
    for i, coin_data in enumerate(coins_data):
        ws.cell(row=next_row + 3, column=3 + i, value=clean_numeric(coin_data['amount']))

    # Fifth row: "Value" in column 2
    ws.cell(row=next_row + 4, column=2, value="Value")

    # Write values for each coin
    for i, coin_data in enumerate(coins_data):
        ws.cell(row=next_row + 4, column=3 + i, value=clean_numeric(coin_data['value']))

    # Set column widths for better readability
    ws.column_dimensions['A'].width = 20  # Date column
    ws.column_dimensions['B'].width = 15  # Labels column
    for i in range(len(coins_data)):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 15

    # Save the workbook
    wb.save(file_name)
    #print(f"Data written to '{file_name}' in worksheet '{ws}' starting at row {next_row}")
    #print(f"Layout: {len(coins_data)} coins written horizontally")

# Example usage function (for testing)
def example_usage():
    """Example of how to use the function"""
    sample_data = [
        {'coin': 'BTC', 'amount': 1.5, 'rate': 45000.50, 'value': 67500.75},
        {'coin': 'ETH', 'amount': 10.2, 'rate': 3200.25, 'value': 32642.55},
        {'coin': 'ADA', 'amount': 1000, 'rate': 0.45, 'value': 450.00}
    ]
    write_horizontal(sample_data, storage)

# Alternative: If you want to keep your original function signature but call it multiple times
def write_single_entry(coin, amount, rate, value):
    """
    Modified version of the original function that works with the horizontal layout
    This version adds a single coin to the existing horizontal layout
    """
    file_name = "tracker.xlsx"

    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if "Warm_History" in wb.sheetnames:
        ws = wb["Warm_History"]
    else:
        ws = wb.create_sheet("Warm_History")
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    def clean_numeric(value):
        if isinstance(value, str):
            return float(value.replace("€", "").replace(",", "").replace("$", ""))
        return float(value)

    # Find the current date row (look for DATE in column 1)
    date_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "DATE":
            date_row = row
            break

    if date_row is None:
        # First entry - create the structure
        date_row = 1
        ws.cell(row=date_row, column=1, value="DATE")
        ws.cell(row=date_row, column=3, value="COINS")
        ws.cell(row=date_row + 1, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=date_row + 2, column=2, value="Rate")
        ws.cell(row=date_row + 3, column=2, value="Amount")
        ws.cell(row=date_row + 4, column=2, value="Value")

    # Find the next available column (first empty column starting from column 3)
    next_col = 3
    while ws.cell(row=date_row + 1, column=next_col).value is not None:
        next_col += 1

    # Write the coin data
    ws.cell(row=date_row + 1, column=next_col, value=coin)  # Coin name
    ws.cell(row=date_row + 2, column=next_col, value=clean_numeric(rate))  # Rate
    ws.cell(row=date_row + 3, column=next_col, value=clean_numeric(amount))  # Amount
    ws.cell(row=date_row + 4, column=next_col, value=clean_numeric(value))  # Value

    # Set column width
    col_letter = get_column_letter(next_col)
    ws.column_dimensions[col_letter].width = 15

    wb.save(file_name)
    print(f"Coin '{coin}' added to '{file_name}' in column {next_col}")






@debug_log
def show_total_assets(root, main_widgets):
    global balances, is_tracker_active
    global warm_value, cold_value, total_assets_value_label
    global updater_job_total, animation_job_id, write_total # <--- Added animation_job_id here
    global total_stock_value, par_demo_mode
    global status_label, btc_label # Make sure btc_label is accessible
    global T_EUR_I, T_EUR_O, T_INVST, T_PL, pl_percentage,btc_price
    global warm_value_var, cold_value_var, total_assets_value_var, total_perc_var, total_crypto_text_var, total_pl_var

    root.title("Total Assets - Crypto Price Tracker V1.5")
    icon_path = os.path.join(os.getcwd(), "crypto", f"MoneyTot.ico")
    root.iconbitmap(icon_path)  # Your .ico file path here
    root.configure(bg=bg_color)

    is_tracker_active = False # This variable isn't directly used to control the `after` loops
    for menu in menubar.children.values():  # Iterate through all menus
        for i in range(menu.index('end') + 1):  # Loop through each item
            menu.entryconfig(i, state="disabled")  # Disable each item

    if par_demo_mode:
        T_EUR_I = 40000
        T_EUR_O = 10000
        T_INVST = T_EUR_I - T_EUR_O
    else:
        file_path = 'tracker.xlsx'
        succes = find_eur_and_get_amounts(file_path)
        if succes:
            print(f"For EUR, the 'Amount deposit' is: {T_EUR_I} and 'Amount withdrawal' is: {T_EUR_O}")
            #T_EUR_I = deposit_value
            #T_EUR_O = abs(withdrawal_value)
            T_INVST = T_EUR_I - T_EUR_O
        else:
            print("❌ No values found")



    C_counter = 0
    C_date = None
    # total_stock_value is already global and initialized, no need to re-initialize here
    # T_PL is already global and initialized, no need to re-initialize here
    if par_demo_mode:
        total_stock_value= 25000
    else:
        if total_stocks == 0:
            try:
                wb = openpyxl.load_workbook('tracker.xlsx')
                ws = wb['Stocks']
                # Get latest row in worksheet
                latest_row = max((cell.row for row in ws.iter_rows() for cell in row if cell.value), default=1)
                C_counter = latest_row
                C_date = ws['A' + str(C_counter)].value
                total_stock_value = ws['B' + str(C_counter)].value
            except FileNotFoundError:
                print("assigned total_stock_value is wrong / file not found or sheet not found")
            except KeyError:
                print("Sheet 'Stocks' or cell 'C1'/'A'/'B' not found in tracker.xlsx")
            except Exception as e:
                print(f"An error occurred while reading Excel: {e}")
        else:
            total_stock_value = total_stocks

    total_pl_var = tk.StringVar(root, value="0.00")

    warm_value_var = tk.StringVar(root, value="0.00")
    cold_value_var = tk.StringVar(root, value="0.00")
    total_assets_value_var = tk.StringVar(root, value="0.00")
    total_perc_var = tk.StringVar(root, value="0.00")
    total_crypto_text_var = tk.StringVar(root, value="0.0)") # Renamed for consistency

    def update_assets():
        global warm_value_var, cold_value_var, total_assets_value_var, pl_percentage, total_perc_var, total_pl_var
        global updater_job_total # Ensure we access the global one to reschedule
        global T_PL, T_INVST, total_stock_value
        global status_label, btc_label # Need to access these to update them

        warm_balances = get_warm_exchange_balance()
        cold_balances = get_cold_storage_balance()

        prices = {}
        all_coins = set(warm_balances.keys()) | set(cold_balances.keys())
        for coin in all_coins:
            prices[coin] = get_crypto_ticker(coin)

        total_warm_value = 0
        if warm_balances:
            total_warm_value = sum(
                (balance['available'] + balance['in_order']) * (
                    prices.get(coin, {}).get('eur_rate', 1) if prices.get(coin) and coin != 'EUR' else 1
                )
                for coin, balance in warm_balances.items()
                if (prices.get(coin) and prices.get(coin).get('eur_rate')) or coin == 'EUR'
            )
            warm_value_var.set(f"€{total_warm_value:.2f}")

        total_cold_value = 0
        if cold_balances:
            total_cold_value = sum(
                amount * (
                    prices.get(coin, {}).get('eur_rate', 1) if prices.get(coin) and coin != 'EUR' else 1
                )
                for coin, amount in cold_balances.items()
                if (prices.get(coin) and prices.get(coin).get('eur_rate')) or coin == 'EUR'
            )
            cold_value_var.set(f"€{total_cold_value:.2f}")

        # Calculate T_PL after both warm and cold values are determined


        T_PL = (total_cold_value + total_warm_value) - T_INVST
        if T_INVST != 0:
            pl_percentage = (((T_PL / T_INVST) * 100) - 100)
        else:
            pl_percentage = 0.00 # Avoid division by zero

        total_pl_var.set(f"€{T_PL:.2f}")
        total_perc_var.set(f"{pl_percentage:.2f} %")
        total_crypto_text_var.set(f"Current Crypto Profit/Loss: ({total_perc_var.get()})")

        total_assets_value = total_warm_value + total_cold_value + total_stock_value
        total_assets_value_var.set(f"€{total_assets_value:.2f}")

        # Re-schedule the next asset update
        updater_job_total = root.after(par_refresh_total, update_assets)

        # Call animate_status after update_assets has run once
        animate_status()


    def animate_status(frame_idx=0, elapsed=0):
        global animation_job_id, status_label, btc_label # Access global for cancellation
        symbols = ["🔄", "🔃"]
        frame_interval = 300
        total_animation_time = 3000

        if status_label is None or not status_label.winfo_exists():
            # If the label doesn't exist, stop trying to animate
            print("Status label does not exist, stopping animation.")
            return

        if elapsed < total_animation_time:
            # Clear btc_label during animation, or update it if needed
            if btc_label and btc_label.winfo_exists():
                 btc_label.config(text="") # Clear it with spaces
            status_label.config(text=f"{symbols[frame_idx % len(symbols)]}", fg=fg_cyan)
            # Store the job ID so it can be cancelled
            animation_job_id = root.after(frame_interval, animate_status, frame_idx + 1, elapsed + frame_interval)
        else:
            if status_label and status_label.winfo_exists():
                btc_val = get_coin_exchange_ticker('BTC-EUR')
                btc_price = btc_val["price"]
                status_label.config(text="✅", fg=fg_ani)
                if btc_label and btc_label.winfo_exists():
                    eur_usd_rate = scrape_eur_usd()

                    try:
                        btc_price_usd = float(btc_price) * eur_usd_rate
                        formatted_price = str(int(round(btc_price_usd, 0)))

                    except ValueError:
                        print("Error: btc_price is not a valid number!")

                    btc_label.config(text=" Current Bitcoin Price: € " + btc_price + " / $ " + formatted_price, fg=fg_color)
            animation_job_id = None # Animation finished, clear the job ID


    def back_to_main():
        global btc_price
        btc_val=get_coin_exchange_ticker('BTC-EUR')
        btc_price=btc_val["price"]
        print(par_write_total)
        if par_write_total:
            write_totals(warm_value_var.get(), cold_value_var.get(), total_stock_value,
                    total_assets_value_var.get(),T_EUR_I, T_EUR_O, T_INVST, total_pl_var.get(), pl_percentage,
                    btc_price)

        root.title("Main - Crypto Price Tracker V1.5")
        icon_path = os.path.join(os.getcwd(), "crypto", f"MoB.ico")
        root.iconbitmap(icon_path)  # Your .ico file path here
        root.configure(bg=bg_color)

        global is_tracker_active, updater_job_total, animation_job_id
        is_tracker_active = True # Set to true if returning to main, assuming main is 'active'
        for menu in menubar.children.values():
            for i in range(menu.index('end') + 1):
                menu.entryconfig(i, state="normal")


        # Cancel the asset update job
        if updater_job_total:
            root.after_cancel(updater_job_total)
            updater_job_total = None
            print("Cancelled updater_job_total")

        # Cancel the animation job
        if animation_job_id:
            root.after_cancel(animation_job_id)
            animation_job_id = None
            print("Cancelled animation_job_id")

        # Destroy all widgets for the current screen
        for widget in root.winfo_children():
            if not isinstance(widget, tk.Menu):
                widget.destroy()

        show_main_screen(root) # Call your main screen function


    # --- Widget Creation ---
    # Destroy all non-menu widgets from previous screen
    for widget in root.winfo_children():
        if not isinstance(widget, tk.Menu):
            widget.pack_forget()
            widget.place_forget()
            widget.destroy()

    # Reset status_label and btc_label references after destroying
    global status_label, btc_label
    status_label = None
    btc_label = None

    root.geometry("700x700")
    root.configure(bg=bg_color)
    if par_demo_mode:
        total_storage_frame = tk.LabelFrame(
            root,
            text="Assets Overview ** Demo **",
            font=("Helvetica", 22, "bold"),
            fg=fg_color,
            bg=bg_color,
            height=600,
            bd=2,
            relief="groove",
            labelanchor="n")
        #title_label = tk.Label(root, text="Assets Overview * Demo Mode *", font=("Helvetica", 20, "bold"), fg=fg_color, bg=bg_color)
    else:
        total_storage_frame = tk.LabelFrame(
            root,
            text="Assets Overview",
            font=("Helvetica", 22, "bold"),
            fg=fg_color,
            bg=bg_color,
            height=600,
            bd=2,
            relief="groove",
            labelanchor="n")
        #title_label = tk.Label(root, text="Assets Overview", font=("Helvetica", 20, "bold"), fg=fg_color, bg=bg_color)

    #total_storage_frame.pack(pady=10, fill="x", padx=20)
    total_storage_frame.pack(fill="both", expand=False, padx=20, pady=10)
    total_storage_frame.pack_propagate(False) # makes sure the height will be kept

    tot_frame = tk.LabelFrame(
        total_storage_frame,
        text="Total Assets",
        font=("Helvetica", 14, "bold"),
        fg=fg_tot_assets,
        bg=bg_color,
        bd=2,
        relief="groove",
        labelanchor="n"
    )
    tot_frame.pack(pady=5, fill="x", padx=20)

    # Warm Storage
    warm_frame = tk.Frame(tot_frame, bg=bg_color)

    warm_frame.pack(pady=3, fill="x")
    warm_label = tk.Label(warm_frame, text="Value Warm Storage:", font=("Helvetica", 14, "bold"), fg="Orange", bg=bg_color, anchor="w")
    warm_label.pack(side="left", padx=(20, 0))
    warm_value = tk.Label(warm_frame, textvariable=warm_value_var, font=("Helvetica", 14,"bold"), fg="Orange", bg=bg_color, anchor="e")
    warm_value.pack(side="right", padx=(0, 20))

    # Cold Storage
    cold_frame = tk.Frame(tot_frame, bg=bg_color)
    cold_frame.pack(pady=3, fill="x")
    cold_label = tk.Label(cold_frame, text="Value Cold Storage:", font=("Helvetica", 14), fg=fg_cold, bg=bg_color, anchor="w")
    cold_label.pack(side="left", padx=(20, 0))
    cold_value = tk.Label(cold_frame, textvariable=cold_value_var, font=("Helvetica", 14), fg=fg_cold, bg=bg_color, anchor="e")
    cold_value.pack(side="right", padx=(0, 20))

    # Stocks
    stock_frame = tk.Frame(tot_frame, bg=bg_color)
    stock_frame.pack(pady=3, fill="x")
    if C_date is not None:
        stock_label = tk.Label(stock_frame, text=f"Value Stocks (last known):", font=("Helvetica", 14), fg="Yellow", bg=bg_color, anchor="w")
    else:
        stock_label = tk.Label(stock_frame, text="Value Stocks:", font=("Helvetica", 14), fg="Yellow", bg=bg_color, anchor="w")
    stock_label.pack(side="left", padx=(20, 0))
    stock_value_label_widget = tk.Label(stock_frame, text=f"€{total_stock_value:.2f}", font=("Helvetica", 14), fg="yellow", bg=bg_color, anchor="e")
    stock_value_label_widget.pack(side="right", padx=(0, 20))

    sep0_frame = tk.Frame(tot_frame, bg=bg_color)
    sep0_frame.pack(pady=5, fill="x")
    separator_width = 140
    separator_thickness = 2 # Changed to 2 for a thinner line
    separator_widget = tk.Frame(sep0_frame, height=separator_thickness, width=separator_width, bg=fg_color)
    separator_widget.pack(side="right", padx=(0, 20))

    total_assets_value_frame = tk.Frame(tot_frame, bg=bg_color)
    total_assets_value_frame.pack(pady=5, fill="x")
    total_assets_label = tk.Label(
    total_assets_value_frame,
        text="Total Assets Value:",
        font=("Helvetica", 14, "bold"),
        fg=fg_color,
        bg=bg_color,
        anchor="w"
    )
    total_assets_label.pack(side="left", padx=(20, 0))
    total_assets_value_label = tk.Label(
    total_assets_value_frame,
    textvariable=total_assets_value_var,
    font=("Helvetica", 14, "bold"),
    fg=fg_color,
    bg=bg_color,
    anchor="e"
    )
    total_assets_value_label.pack(side="right", padx=(0, 20))


    # --- STATUS AND BTC LABELS ---
    # Create these labels BEFORE calling update_assets or animate_status,
    # as they need to exist for those functions to configure them.
    status_label = tk.Label(root, text="", font=("Helvetica", 16), fg=fg_cyan, bg=bg_color, anchor="sw")
    status_label.place(x=20, y=660) # Adjust y position as needed, relative to root size
    status_label.lift()


    btc_label = tk.Label(root, text="", font=("Helvetica", 12), fg=fg_color, bg=bg_color, anchor="sw")
    btc_label.place(x=195, y=660) # Adjust y position as needed
    btc_label.lift()

    sep_crypto_frame = tk.Frame(total_storage_frame, bg=bg_color)
    sep_crypto_frame.pack(pady=10, fill="x")
    sep_crypto_label = tk.Label(sep_crypto_frame, text="", font=("Helvetica", 14, "underline"), fg="lightgray", bg=bg_color, anchor="w")
    sep_crypto_label.pack(side="left", padx=(20, 0))


    # LabelFrame for Crypto Investment
    crypto_frame = tk.LabelFrame(
    total_storage_frame,
    text="Crypto Investment",
    font=("Helvetica", 14, "bold"),
    fg=fg_tot_assets,
    bg=bg_color,
    bd=2,
    relief="groove",
    labelanchor="n"
)
    crypto_frame.pack(pady=10, fill="x", padx=20)

    # EUR Investment In
    total_invest_frame = tk.Frame(crypto_frame, bg=bg_color)
    total_invest_frame.pack(pady=3, fill="x")
    total_invest_label = tk.Label(total_invest_frame, text="EUR Investment in", font=("Helvetica", 14, "bold"), fg=fg_day, bg=bg_color, anchor="w")
    total_invest_label.pack(side="left", padx=(20, 0))
    total_invest_value_label = tk.Label(total_invest_frame, text=f"€{T_EUR_I:.2f}", font=("Helvetica", 14, "bold"), fg=fg_day, bg=bg_color, anchor="e")
    total_invest_value_label.pack(side="right", padx=(0, 20))

    # EUR Investment Out
    total_EUR_out_frame = tk.Frame(crypto_frame, bg=bg_color)
    total_EUR_out_frame.pack(pady=3, fill="x")
    total_EUR_out_label = tk.Label(total_EUR_out_frame, text="EUR Investment out", font=("Helvetica", 14, "bold"), fg="lightyellow", bg=bg_color, anchor="w")
    total_EUR_out_label.pack(side="left", padx=(20, 0))
    total_EUR_out_value_label = tk.Label(total_EUR_out_frame, text=f"€{T_EUR_O:.2f}", font=("Helvetica", 14, "bold"), fg="lightyellow", bg=bg_color, anchor="e")
    total_EUR_out_value_label.pack(side="right", padx=(0, 20))

    # Total Investment
    total_current_frame = tk.Frame(crypto_frame, bg=bg_color)
    total_current_frame.pack(pady=3, fill="x")
    total_current_label = tk.Label(total_current_frame, text="Total Investment", font=("Helvetica", 14, "bold"), fg=fg_day, bg=bg_color, anchor="w")
    total_current_label.pack(side="left", padx=(20, 0))
    total_current_value_label = tk.Label(total_current_frame, text=f"€{T_INVST:.2f}", font=("Helvetica", 14, "bold"), fg=fg_day, bg=bg_color, anchor="e")
    total_current_value_label.pack(side="right", padx=(0, 20))

    separator_inner = tk.Frame(crypto_frame, height=1, width=separator_width, bg=fg_color)
    separator_inner.pack(pady=5, anchor="e", padx=(0, 20))



    # Profit/Loss
    total_crypto_profit_loss_frame = tk.Frame(crypto_frame, bg=bg_color)
    total_crypto_profit_loss_frame.pack(pady=3, fill="x")
    total_crypto_label_pl = tk.Label(
        total_crypto_profit_loss_frame,
        textvariable=total_crypto_text_var,
        font=("Helvetica", 14, "bold"),
        fg=fg_tot_crypto,
        bg=bg_color,
        anchor="w"
        )
    total_crypto_label_pl.pack(side="left", padx=(20, 0))
    total_crypto_value_label_pl = tk.Label(
            total_crypto_profit_loss_frame,
            textvariable=total_pl_var,
            font=("Helvetica", 14, "bold"),
            fg=fg_tot_crypto,
            bg=bg_color,
            anchor="e"
            )
    total_crypto_value_label_pl.pack(side="right", padx=(0, 20))
    # Back Button
    try:
        icon_path = os.path.join(os.getcwd(), "crypto", f"back_blue.png")
        img = Image.open(icon_path).resize((20, 20))

        img = img.resize((20, 20))
        img_tk = ImageTk.PhotoImage(img)
    except FileNotFoundError:
        print("back_blue.png not found. Using text button instead.")
        img_tk = None # No image available

    # Mouse pointer hover for back_button


    back_button = Button(
        root,
        image=img_tk,
        command=back_to_main,
        bg="grey",
        activebackground="forestgreen",
        highlightbackground="white",
        bd=5,
        relief="raised",
        width = 35
        )
    back_button.bind("<Enter>", on_hover)
    back_button.bind("<Leave>", on_leave)
    back_button.place(x=630, y=660)
    if img_tk: # Only assign if image was loaded
        back_button.image = img_tk


    # Start the initial update and animation
    update_assets() # This will call animate_status internally for the first time






def call_aggr_window():
    bg="black"
    fg="white"
    window = webview.create_window("Tracker Live View AGGR", "https://www.aggr.trade/remr")
    webview.start()
    return window.evaluate_js("document.title")



def call_mempool_window():
    icon_path = os.path.join(os.getcwd(), "crypto", f"mem.ico")
    icon_filename=(icon_path)
    html_file="https://mempool.space"
    window = webview.create_window(
        "Tracker Live View Mempool",
        html_file,
        frameless=False,
        transparent=True,
        confirm_close=True
    )
    webview.start()
    #window.destroy()
    #root.destroy()
    return window.evaluate_js("document.title")


def call_user_window(name,url):
    window = webview.create_window("Tracker Crypto Sentiment"+name, url)
    webview.start()
    return window.evaluate_js("document.title")


def call_csv_window():
    csv_path = os.path.join(os.path.dirname(__file__), "calcpiv_module.py")
    subprocess.Popen(["python", csv_path])

def call_config_tracker():
    config_path = os.path.join(os.path.dirname(__file__), "config_tracker_module.py")
    print(f"Launching config tracker: {config_path}")

    # Launch the subprocess and store the Popen object
    process = subprocess.Popen(["python", config_path])

    # Wait for the subprocess to complete before continuing
    print("Waiting for config tracker to finish...")
    process.wait() # This line makes your current script pause

    # These lines will only execute AFTER config_tracker_module.py has finished
    print("Config tracker finished. Returning from call_config_tracker.")
    restart_program()


def call_fng():
    global is_tracker_active
    is_tracker_active = False


    config_path = os.path.join(os.path.dirname(__file__), "fng_module.py")
    #subprocess.Popen(["python", config_path])
    subprocess.run(["python", config_path])
    is_tracker_active=True

def call_show_about():
    config_path = os.path.join(os.path.dirname(__file__), "show_readme_module.py")
    subprocess.Popen(["python", config_path])


def get_coin_id(symbol):
    """Search for the correct CoinGecko ID using the symbol."""
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {"vs_currency": "usd", "order": "market_cap_desc", "per_page": 250, "page": 1}
    try:
        response = requests.get(url, params=params).json()
        time.sleep(0.5)  # Add a 0.5-second delay
        if isinstance(response, list):
            for coin in response:
                if "symbol" in coin and coin["symbol"].lower() == symbol.lower():
                    return coin["id"]
        else:
            print(f"Warning: Unexpected response format in get_coin_id: {response}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error in get_coin_id: {e}")
        return None
    return None

from PIL import Image, ImageTk
import os


def get_coin_icon(coin_symbol):
    """Search the icon (.png file) for the coin"""
    icon_path = os.path.join(os.getcwd(), "crypto", "ico", "32", f"{coin_symbol.lower()}.png")  # filename dynamic

    # Check if directory excists
    if os.path.exists(icon_path):
        try:
            image = Image.open(icon_path)
            image = image.resize((20, 20))  # Optional: adjust size
            return ImageTk.PhotoImage(image)
        except Exception as e:
            print(f"Error loading Icon: {e}")
            return None
    else:
        coin_notfound="generic"
        icon_path = os.path.join(os.getcwd(), "crypto", "ico", "32", f"{coin_notfound.lower()}.png")
        image = Image.open(icon_path)
        image = image.resize((20, 20))  # Optional: adjust size
        print(f"Icon not found for: {coin_symbol}")
        return ImageTk.PhotoImage(image)


def get_coin_icon_main(coin_symbol):
    """Search the icon (.png file) for the coin"""
    icon_path = os.path.join(os.getcwd(), "crypto", "ico", "32", f"{coin_symbol.lower()}.png")  # # filename dynamic

    # Controleer of het bestand bestaat
    if os.path.exists(icon_path):
        try:
            image = Image.open(icon_path)
            image = image.resize((35, 35))  # Optional: adjust size
            return ImageTk.PhotoImage(image)
        except Exception as e:
            print(f"Error loading Icon: {e}")
            return None
    else:
        coin_notfound="generic"
        icon_path = os.path.join(os.getcwd(), "crypto", "ico", "32", f"{coin_notfound.lower()}.png")
        image = Image.open(icon_path)
        image = image.resize((20, 20))  # Optional: adjust size
        print(f"Icon not found for: {coin_symbol}")
        return ImageTk.PhotoImage(image)




# Get All Time high Value for the coin

def get_ath(symbol):
    """Fetch ATH prices in both USD and EUR and return as separate values."""
    coin_id = get_coin_id(symbol)

    if not coin_id:
        return None, None

    url = f"https://api.coingecko.com/api/v3/coins/{coin_id}"
    try:
        response = requests.get(url).json()
        time.sleep(2.1)  # Add a 0.5-second delay
        if "market_data" in response and "ath" in response["market_data"] and "usd" in response["market_data"]["ath"] and "eur" in response["market_data"]["ath"]:
            ath_usd = response["market_data"]["ath"]["usd"]
            ath_eur = response["market_data"]["ath"]["eur"]
            return ath_usd, ath_eur
        else:
            print(f"Warning: Could not retrieve ATH data for {symbol} (ID: {coin_id}). Response: {response}")
            return None, None
    except requests.exceptions.RequestException as e:
        print(f"Error in get_ath: {e}")
        return None, None


def open_excel_file(excel_filepath):
    import os
    import platform
    import subprocess
    current_directory = os.getcwd()
    try:
        if platform.system() == "Windows":
            os.startfile(excel_filepath)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", excel_filepath])
        elif platform.system() == "Linux":
            subprocess.run(["xdg-open", excel_filepath])
        else:
            print(f"Unsupported operating system for opening files: {platform.system()}")
    except FileNotFoundError:
        print(f"Error: File not found at {excel_filepath}")
    except Exception as e:
        print(f"An error occurred: {e}")


def init_excel(filename="tracker.xlsx"):
    """
    Creates a new Excel workbook with the worksheets "credentials" and "cold_storage".
    Adds some example data to each worksheet and saves the file.
    Displays a warning screen if the file already exists.

    Args:
    filename (str, optional): The name under which the Excel file
    should be saved. Defaults to "tracker.xlsx".
    """
    import os
    if os.path.exists(filename):
        root = tk.Tk()
        root.resizable(False, False)
        root.withdraw()  # Verberg het hoofdvenster

        def on_yes():
            root.destroy()
            try_create_and_save(filename)

        def on_no():
            root.destroy()
            print(f"Creation of '{filename}' has been canceled")

        message_window = tk.Toplevel(root)
        message_window.title("Warning")
        message_label = tk.Label(message_window, text="Are you sure?", fg="red", font=("Arial", 12))
        message_label.pack(pady=10)

        yes_button = tk.Button(message_window, text="Yes", command=on_yes, width=8, height=1)
        yes_button.pack(side=tk.LEFT, padx=5, pady=10)

        no_button = tk.Button(message_window, text="No", command=on_no, width=8, height=1)
        no_button.pack(side=tk.LEFT, padx=5, pady=10)

        root.mainloop()
    else:
        try_create_and_save(filename)




def try_create_and_save(filename):
    """
    Tries to create and save the Excel file with specified cell colors.
    """
    try:
        # Create a new empty workbook
        workbook = Workbook()

        # Remove the default worksheet "Sheet"
        std_sheet = workbook.active
        workbook.remove(std_sheet)


        # Color settings for the worksheets

        lightorange_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
        lightblue_fill = PatternFill(start_color="C5E7ED", end_color="C5E7ED", fill_type="solid")
        lightyellow_fill = PatternFill(start_color="F6FAD2", end_color="F6FAD2", fill_type="solid")
        lightgreen_fill = PatternFill(start_color="CCFCE7", end_color="CCFCE7", fill_type="solid")


        # Create the "credentials" worksheet
        credentials_sheet = workbook.create_sheet("credentials")
        credentials_sheet['A1'] = "Warm Storage"
        credentials_sheet['B1'] = "Value"
        credentials_sheet['A2'] = "Exchange Name"
        credentials_sheet['A3'] = "API Key"
        credentials_sheet['A4'] = "Secret Key"
        credentials_sheet['A5'] = "URL"
        credentials_sheet['C1'] = ">>> Essential for tracker <<<"

        # Set the width of column A for the "credentials" worksheet
        credentials_sheet.column_dimensions['A'].width = 40
        credentials_sheet.column_dimensions['B'].width = 80

        # Set the background color for the header row in "credentials" to light orange
        lightorange_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
        lightblue_fill = PatternFill(start_color="C5E7ED", end_color="C5E7ED", fill_type="solid")
        lightyellow_fill = PatternFill(start_color="F6FAD2", end_color="F6FAD2", fill_type="solid")
        lightgreen_fill = PatternFill(start_color="CCFCE7", end_color="CCFCE7", fill_type="solid")
        credentials_sheet['A1'].fill = lightorange_fill
        credentials_sheet['B1'].fill = lightorange_fill
        credentials_sheet['A2'].fill = lightyellow_fill
        credentials_sheet['A3'].fill = lightyellow_fill
        credentials_sheet['A4'].fill = lightyellow_fill
        credentials_sheet['A5'].fill = lightyellow_fill
        credentials_sheet['B2'].fill = lightgreen_fill
        credentials_sheet['B3'].fill = lightgreen_fill
        credentials_sheet['B4'].fill = lightgreen_fill
        credentials_sheet['B5'].fill = lightgreen_fill

        # Formatting for the title "Exchange"
        title_cell = credentials_sheet['A1']
        title_cell.font = Font(bold=True, size=16)
        title_cell = credentials_sheet['B1']
        title_cell.font = Font(bold=True, size=16)


        # Create the "Cold_Storage" worksheet
        cold_storage_sheet = workbook.create_sheet("Cold_Storage")
        cold_storage_sheet['A1'] = "COLD STORAGE"
        cold_storage_sheet['A2'] = "Coin"
        cold_storage_sheet['B2'] = "Amount"
        cold_storage_sheet['A3'] = "BTC"
        cold_storage_sheet['B3'] = "0.5"
        cold_storage_sheet['A4'] = "ETH"
        cold_storage_sheet['B4'] = "2.0"
        cold_storage_sheet['C1'] = ">>> Essential for tracker <<<"

        # Formatting for the title "COLD STORAGE"
        title_cell = cold_storage_sheet['A1']
        title_cell.font = Font(bold=True, size=16)

        # Formatting for the title "Exchange"
        title_cell = credentials_sheet['A1']
        title_cell.font = Font(bold=True, size=16)

        # Formatting for the headers of the table in "Cold_Storage" to light blue
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        cold_storage_sheet['A1'].fill = light_blue_fill
        cold_storage_sheet['B1'].fill = light_blue_fill
        cold_storage_sheet['A2'].fill = lightblue_fill
        cold_storage_sheet['B2'].fill = lightblue_fill

         # Fill the range A3 to A50 with lightyellow also B3 to B50

        for row in range(3, 51):  # Iterate through rows 3 to 50 (inclusive)
            cell = cold_storage_sheet[f'A{row}']
            cell.fill = lightyellow_fill
            cell = cold_storage_sheet[f'B{row}']
            cell.fill = lightgreen_fill


        # Set the width of column A and B for the "Cold_Storage" worksheet
        cold_storage_sheet.column_dimensions['A'].width = 40
        cold_storage_sheet.column_dimensions['B'].width = 10

        # Save the workbook under the specified filename
        workbook.save(filename)

        print(f"Excel file '{filename}' successfully created and saved with specified colors.")

    except Exception as e:
        print(f"An error occurred while creating and saving the Excel file: {e}")


def show_main_screen(root):

    global is_tracker_active, after_id, main_widgets, available_coins,default_coin, coin_list

    is_tracker_active = True
    print("Show main screen")

    # Hide all other widgets and reset the screen
    for widget in root.winfo_children():
        widget.pack_forget()
        widget.place_forget()
        widget.grid_forget()

    # Restore the root window size
    root.geometry("600x300")

    # Recreate the main screen widgets
    header_frame = tk.Frame(root, bg=bg_color)
    header_frame.pack(pady=(10, 5))
    main_labels = {
        'header_white': tk.Label(header_frame, text="Current: ", font=("Helvetica", 22, "bold"), fg=fg_color, bg=bg_color),
        'header_orange': tk.Label(header_frame, text="", font=("Helvetica", 22, "bold"), fg="orange", bg=bg_color),
        'footer_frame': tk.Frame(root, bg=bg_color)
    }
    main_labels['header_white'].pack(side="left")
    main_labels['header_orange'].pack(side="left")

    # Shared frame for EUR and USD to align perfectly
    main_labels['rates_frame'] = tk.Frame(root, bg=bg_color)
    main_labels['rates_frame'].pack(pady=5)

    # EUR row
    main_labels['eur_text'] = tk.Label(main_labels['rates_frame'], text="EUR:", font=("Helvetica", 16), fg=fg_color, bg=bg_color, anchor="e", width=5)
    main_labels['eur_text'].grid(row=0, column=0, sticky="e")
    main_labels['eur_value'] = tk.Label(main_labels['rates_frame'], text="Loading...", font=("Helvetica", 16), fg=fg_color, bg=bg_color)
    main_labels['eur_value'].grid(row=0, column=1, sticky="w")

    # USD row
    main_labels['usd_text'] = tk.Label(main_labels['rates_frame'], text="USD:", font=("Helvetica", 16), fg=fg_color, bg=bg_color, anchor="e", width=5)
    main_labels['usd_text'].grid(row=1, column=0, sticky="e")
    main_labels['usd_value'] = tk.Label(main_labels['rates_frame'], text="Loading...", font=("Helvetica", 16), fg=fg_color, bg=bg_color)
    main_labels['usd_value'].grid(row=1, column=1, sticky="w")

    main_labels['footer_frame'].pack(pady=(5, 10))
    main_labels['footer_text'] = tk.Label(main_labels['footer_frame'], text="Updated:", font=("Helvetica", 16), fg=fg_color, bg=bg_color)
    main_labels['footer_text'].pack(side="left")
    main_labels['footer_date'] = tk.Label(main_labels['footer_frame'], text="Loading...", font=("Helvetica", 16), fg="yellow", bg=bg_color)
    main_labels['footer_date'].pack(side="left")

    # Extra label for the exchange rate (EUR/USD)
    eur_usd_rate = scrape_eur_usd()
    exchange_rate_text = "Rate: Loading..."
    if eur_usd_rate is not None:
        usd_eur_rate = 1 / eur_usd_rate if eur_usd_rate != 0 else "N/A"
        exchange_rate_text = f"Rate: €{eur_usd_rate:.4f} / ${usd_eur_rate:.4f}"
        print("Eur rate", eur_usd_rate)

    exchange_rate_label = tk.Label(root, text=exchange_rate_text, font=("Helvetica", 10), fg=fg_cyan, bg=bg_color, anchor="se")
    exchange_rate_label.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)

    #Extra for Ath
    ath_label_text = tk.StringVar()
    ath_label_text.set("Loading ATH...") # Set initial text
    ath_label = tk.Label(
        root, # Parent is now root
        textvariable=ath_label_text,
        font="Helvetica, 10",
        fg=fg_cyan,
        bg=bg_color,
        anchor="sw"
    )
    ath_cache = ath_label
    ath_label.place(relx=0.0, rely=1.0, anchor="sw", x=10, y=-10)


    # Dropdown menu for selecting the cryptocurrency
    coins_dropdown = ttk.Combobox(root, textvariable=selected_coin, values=available_coins, state="readonly")

    if default_coin in coin_list:
        default_coin_index = available_coins.index(default_coin)
    elif coin_list: # If default coin not found, but list is not empty, select the first one
        coins_dropdown.set(available_coins[0])

    coins_dropdown.pack(pady=10)
    coins_dropdown.set(available_coins[default_coin_index])
    coins_dropdown.bind("<<ComboboxSelected>>", lambda event: update_gui(root, main_widgets))

    # Save references for later use
    main_widgets = {
        'header_frame': header_frame,
        'header_white': main_labels['header_white'],
        'header_orange': main_labels['header_orange'],
        'rates_frame': main_labels['rates_frame'],
        'eur_text': main_labels['eur_text'],
        'eur_value': main_labels['eur_value'],
        'usd_text': main_labels['usd_text'],
        'usd_value': main_labels['usd_value'],
        'footer_frame': main_labels['footer_frame'],
        'footer_text': main_labels['footer_text'],
        'footer_date': main_labels['footer_date'],
        'coins_dropdown': coins_dropdown,
        'exchange_rate_label': exchange_rate_label,
        'ath_label' : ath_label,
        'ath_label_text': ath_label_text, # Add the StringVar here
        #'ath_label_text': ath_label_text
    }

    # Cancel any existing update timers before starting a new one
    if after_id:
        root.after_cancel(after_id)

    # Start the periodic GUI update for the main screen
    after_id = root.after(par_refresh_main, update_gui, root, main_widgets)

    # Perform an immediate update of the values
    update_gui(root, main_widgets)





@debug_log
def main(root=None):
    global selected_coin, is_tracker_active, coin_symbols, menubar, main_widgets
    global ath_price_eu, ath_price_usd, ath_price_str, ath_coin_symbol
    global fg_color,bg_color, fg_cold, fg_cyan, fg_day, fg_ani,fg_tot_assets
    global fg_tot_crypto, fg_tot_storage
    global par_write_total, par_write_warm, par_write_cold, par_demo_mode
    global par_refresh_main,par_refresh_warm, par_refresh_cold, par_refresh_total
    global available_coins,default_coin, coin_list




    filename='tracker.xlsx'
    if os.path.exists(filename):
        print("\n✅ Tracker.xlsx found")
    else:
        print("\n❌ No Tracker.xlsx found.")
        print("\n ..... Going to create it")
        init_excel(filename)
        open_excel_file(filename)
    if os.path.exists('fng_module.py') and os.path.exists('crypto_ticker_module.py'):
        print("\n✅ fng_module.py and crypto_ticker_module.py found")
    else:
        print("\n❌ No fng_module.py found or crypto_ticker_module.py")
        print("\n❌ Fear & Greed will not work")
    if os.path.exists('calcpiv_module.py'):
        print("\n✅ calcpiv_module.py found")
    else:
        print("\n❌ No calcpiv_module.py found.")
        print("\n❌ load & Calculate CSV will not work")

    if os.path.exists('show_readme_module.py'):
            print("\n✅ show_readme_module.py found")

    #
    # Read tracker.cfg and set VARIABLES par_refresh, par_user_name,
    # par_user_url,par_write
    #

    app_settings = load_app_settings()


    if app_settings:
        print("\n✅ Configuration loaded successfully!")

        darkmod = app_settings['dark_mode']
        if darkmod is True:
            print(f"   - Dark Mode Enabled: {app_settings['dark_mode']}")
        else:
            print(f"   - Light Mode Enabled: True")
        par_user_name1 = app_settings.get('Name1', 'Not set')

        par_user_url1=app_settings.get('url1', 'Not set')
        par_user_name2 = app_settings.get('Name2', 'Not set')
        par_user_url2=app_settings.get('url2', 'Not set')
        par_user_name3 = app_settings.get('Name3', 'Not set')
        par_user_url3=app_settings.get('url3', 'Not set')
        par_write_warm = app_settings.get('write_warm')
        par_write_cold = app_settings.get('write_cold')
        par_write_total = app_settings.get('write_total')
        par_refresh_main = app_settings.get('refresh_main')
        par_refresh_warm = app_settings.get('refresh_warm')
        par_refresh_cold = app_settings.get('refresh_cold')
        par_refresh_total = app_settings.get('refresh_total')
        #
        # Multiply RefreshRate * 1000 to obtain seconds
        #
        par_refresh_main = par_refresh_main * 1000
        par_refresh_warm = par_refresh_warm * 1000
        par_refresh_cold = par_refresh_cold * 1000
        par_refresh_total = par_refresh_total * 1000

        par_debug_mode = app_settings.get('debug_mode')
        par_demo_mode = app_settings.get('demo_mode')


        log_level = logging.DEBUG if app_settings['debug_mode'] else logging.INFO
        handlers = [
            logging.StreamHandler(sys.stdout) # Logging to console
            ]



        # check if debug_mode is activated and set the logging correctly

        if app_settings['debug_mode']:
            print("\n🐞 Debug mode is active. Logging to 'tracker_log.log'.")
            handlers.append(logging.FileHandler(
                "tracker_log.log",
                mode="w" # rewrite
                ))

            logging.basicConfig(
                encoding="utf-8",
                level=log_level,
                format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
                handlers=handlers
                )

        if par_demo_mode:
            print("\n🐞 Demo mode is active. No real data.")




    # parameter setting dark/light mode
    if darkmod is False:
        bg_color="lightgray"
        fg_color="black"
        fg_cyan="blue"
        fg_cold="blue"
        fg_day= "green"
        fg_ani="green"
        fg_tot_assets="black"
        fg_tot_crypto="black"
        fg_tot_storage='black'

    else:
        bg_color="black"
        fg_color="white"
        fg_cyan="cyan"
        fg_cold="lightblue"
        fg_day="lightgray"
        fg_ani="lightgreen"
        fg_tot_assets="lightgray"
        fg_tot_crypto="lightgreen"
        fg_tot_storage='orange'

    ath_price_eur = 0
    ath_price_usd = 0

    initial_width = 600
    initial_height = 300
    title_font = ("Helvetica", 22, "bold")
    small_font = ("Helvetica", 10)
    # Set default in Combobox
    default_coin = 'BTC'


    # Build the list with owned coins. Will be used in dropdown main screen
    # 1. Get the balances
    warm_balances = get_warm_exchange_balance()
    cold_balances = get_cold_storage_balance()

    # 2. Create a list of unique coin names
    all_coins = set() # Using a set to automatically handle duplicates

    # Add coins from warm balances
    for coin in warm_balances.keys():
        all_coins.add(coin)
        #all_coins.add(f"{coin} ({coin_symbols.get(coin, 'N/A')})")
        #print(all_coins)

    # Add coins from cold balances
    for coin in cold_balances.keys():
        all_coins.add(coin)

    # Convert the set to a sorted list for the dropdown
    coin_list = sorted(list(all_coins))
    available_coins = coin_list



    #print(f"main() called with root: {root}")
    if root is None:
        try:
            root = tk.Tk()
            root.resizable(False, False)
            root.title("Main - Crypto Price Tracker V1.5")
            icon_path = os.path.join(os.getcwd(), "crypto", f"MoB.ico")
            root.iconbitmap(icon_path)  # Your .ico file path here
            root.configure(bg=bg_color)
            menubar = Menu(root)
            #print(root)
            #print(f"New root created: {root}, menubar: {menubar}")

            def call_show_warm_storage():
                global main_widgets
                show_warm_storage(root)

            def call_show_cold_storage():
                global main_widgets
                show_cold_storage(root, main_widgets)

            def call_show_combined_storage():
                global main_widgets
                #show_cold_combined(root, main_widgets)
                show_combined_storage(root, main_widgets)

            def call_show_total_assets():
                global main_widgets
                show_total_assets(root, main_widgets)


            def call_load_csv_calculate():
                global main_widgets
                load_csv_calculate(root, main_widgets)


            # Options Menu
            options_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Options", menu=options_menu)
            options_menu.add_command(label="Warm Storage", command=call_show_warm_storage)
            options_menu.add_command(label="Cold Storage", command=call_show_cold_storage)
            options_menu.add_command(label="Crypto Storage", command=call_show_combined_storage)
            options_menu.add_command(label="Input Stocks", command=lambda: set_total_stocks(root))
            options_menu.add_command(label="Total Assets", command=call_show_total_assets)

            # External menu
            external_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Crypto Sentiment", menu=external_menu)
            external_menu.add_command(label="Fear and Greed", command=call_fng)
            external_menu.add_command(label="AGGR Live View", command=call_aggr_window)

            #external_menu.add_command(par_user_name(x), par_user_url(x))
            external_menu.add_command(label=par_user_name1, command=lambda: call_user_window(par_user_name1,par_user_url1))
            external_menu.add_command(label=par_user_name2, command=lambda: call_user_window(par_user_name2,par_user_url2))
            external_menu.add_command(label=par_user_name3, command=lambda: call_user_window(par_user_name3,par_user_url3))

            # History Menu / CSV menu
            History_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="CSV Data", menu=History_menu)
            History_menu.add_command(label="Load & Calculate", command=call_csv_window)


            # Config Menu
            config_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="Config", menu=config_menu)
            config_menu.add_command(label="Parameters", command=call_config_tracker)
            config_menu.add_command(label="Open Excel", command=lambda: open_excel_file('tracker.xlsx'))
            config_menu.add_command(label="Init Excel", command=init_excel)



            # About Menu
            About_menu = Menu(menubar, tearoff=0)
            menubar.add_cascade(label="About", menu=About_menu)
            About_menu.add_command(label="About", command=call_show_about)

            root.config(menu=menubar)
            #print(f"Root config('menu') after creation: {root.config('menu')}")

        except Exception as e:
            logging.error(f"Failure initializing root: {e}")
            print(f"Failure initializing root: {e}")
            return
    else:
        print(f"Excisting root accepted: {root}, current menu config: {root.config('menu')}")
        print(f"Widgets in root by return: {root.winfo_children()}")

    selected_coin = tk.StringVar(root)


    try:
        root.geometry(f"{initial_width}x{initial_height}")

        header_frame = tk.Frame(root, bg=bg_color)
        header_frame.pack(pady=(10, 5))
        main_labels = {
            'header_white': tk.Label(header_frame, text="Current: ", font=("Helvetica", 22, "bold"), fg=fg_color, bg=bg_color),
            'header_orange': tk.Label(header_frame, text="", font=("Helvetica", 22, "bold"), fg="orange", bg=bg_color),
            'footer_frame': tk.Frame(root, bg=bg_color)
        }
        main_labels['header_white'].pack(side="left")
        main_labels['header_orange'].pack(side="left")

        # One frame for EUR USD to align perfectly
        main_labels['rates_frame'] = tk.Frame(root, bg=bg_color)
        main_labels['rates_frame'].pack(pady=5)

        # EUR Line
        main_labels['eur_text'] = tk.Label(main_labels['rates_frame'], text="EUR:", font=title_font, fg=fg_color, bg=bg_color, anchor="e", width=5)
        main_labels['eur_text'].grid(row=0, column=0, sticky="e")
        main_labels['eur_value'] = tk.Label(main_labels['rates_frame'], text="Loading...", font=title_font, fg=fg_color, bg=bg_color)
        main_labels['eur_value'].grid(row=0, column=1, sticky="w")

        # USD Line
        main_labels['usd_text'] = tk.Label(main_labels['rates_frame'], text="USD:", font=title_font, fg=fg_color, bg=bg_color, anchor="e", width=5)
        main_labels['usd_text'].grid(row=1, column=0, sticky="e")
        main_labels['usd_value'] = tk.Label(main_labels['rates_frame'], text="Loading...", font=title_font, fg=fg_color, bg=bg_color)
        main_labels['usd_value'].grid(row=1, column=1, sticky="w")


        main_labels['footer_frame'].pack(pady=(5, 10))
        main_labels['footer_text'] = tk.Label(main_labels['footer_frame'], text="Updated:", font=("Helvetica", 16), fg=fg_color, bg=bg_color)
        main_labels['footer_text'].pack(side="left")
        main_labels['footer_date'] = tk.Label(main_labels['footer_frame'], text="Loading...", font=("Helvetica", 16), fg="yellow", bg=bg_color)
        main_labels['footer_date'].pack(side="left")


        # Create and pack the bottom left frame and ATH label
        ath_label_text = tk.StringVar()
        ath_label_text.set("Loading ATH...") # Set initial text
        ath_label = tk.Label(
            root, # Parent is now root
            textvariable=ath_label_text,
            font=small_font,
            fg=fg_cyan,
            bg=bg_color,
            anchor="sw"
        )
        ath_cache = ath_label
        ath_label.place(relx=0.0, rely=1.0, anchor="sw", x=10, y=-10)


        eur_usd_rate = scrape_eur_usd()

        exchange_rate_text = "Rate: Loading..."
        if eur_usd_rate is not None:
            usd_eur_rate = 1 / eur_usd_rate if eur_usd_rate != 0 else "N/A"
            exchange_rate_text = f"Rate: €{eur_usd_rate:.4f} / ${usd_eur_rate:.4f}"


        exchange_rate_label = tk.Label(root, text=exchange_rate_text, font=("Helvetica", 10), fg=fg_cyan, bg=bg_color, anchor="se")
        exchange_rate_label.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)
        ath_label.place(relx=0.0, rely=1.0, anchor="sw", x=10, y=-10)
        #print("ATH on screen")
        #print(ath_cache)
        #print(" before ",ath_price_eur)




        combobox_height = min(len(coin_list), 5) # Show up to 15 items, or fewer if list is shorter

        coins_dropdown = ttk.Combobox(root, textvariable=selected_coin, values=coin_list, state="readonly", height=combobox_height)
        coins_dropdown.pack()
        #coins_dropdown.set(available_coins[0])

        #available_coins = coin_list
        if default_coin in coin_list:
            default_coin_index = coin_list.index(default_coin)
            coins_dropdown.set(coin_list[default_coin_index])

        elif coin_list: # If default coin not found, but list is not empty, select the first one
            coins_dropdown.set(coin_list[0])

        coins_dropdown.bind("<<ComboboxSelected>>", lambda event: update_gui(root, main_widgets))
        #print(selected_coin.get())
        #print(ath_price_eur, ath_price_usd)
        #print("3e Coin printed")

        global main_widgets
        main_widgets = {
            'header_frame': header_frame,
            'header_white': main_labels['header_white'],
            'header_orange': main_labels['header_orange'],
            'rates_frame': main_labels['rates_frame'],
            'eur_text': main_labels['eur_text'],
            'eur_value': main_labels['eur_value'],
            'usd_text': main_labels['usd_text'],
            'usd_value': main_labels['usd_value'],
            'footer_frame': main_labels['footer_frame'],
            'footer_text': main_labels['footer_text'],
            'footer_date': main_labels['footer_date'],
            'coins_dropdown': coins_dropdown,
            'exchange_rate_label': exchange_rate_label,
            'ath_label' : ath_label,
            'ath_label_text': ath_label_text, # Add the StringVar here
            #'ath_label_text': ath_label_text
        }

        def on_closing():
            global is_tracker_active, after_id
            is_tracker_active = False
            if after_id:
                root.after_cancel(after_id)
            root.destroy()

        root.protocol("WM_DELETE_WINDOW", on_closing)

        is_tracker_active = True
        #root.after(par_refresh_main, update_gui, root, main_widgets)
        root.after(1000, update_gui, root, main_widgets)

    except Exception as e:
        logging.error(f"General error in main after initialisation): {e}")
        print(f"General error in main after initialisation): {e}")

    root.mainloop()

if __name__ == "__main__":
    main()
