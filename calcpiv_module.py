import os
import sys
import csv
import time
import json
import hmac
import hashlib
import logging
import requests
import openpyxl
import configparser
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
import platform
import os
import pandas as pd
import logging

# Only import win32com if on Windows
if platform.system() == "Windows":
    try:
        import win32com.client
    except ImportError:
        logging.warning("win32com.client not available. Real PivotTables will be disabled.")

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("crypto_tracker.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

CONFIG_FILE = "tracker.cfg"

# Load API credentials from Excel
try:
    ws = openpyxl.load_workbook('tracker.xlsx')['Credentials']
    Warm_API_Name = ws.cell(row=2, column=2).value
    WARM_API_KEY = ws.cell(row=3, column=2).value
    WARM_API_SECRET = ws.cell(row=4, column=2).value
    WARM_API_URL = ws.cell(row=5, column=2).value
except Exception as e:
    logging.error(f"Error loading API credentials from tracker.xlsx: {e}")
    sys.exit()

# Bitvavo API authentication
def create_signature(ts, method, endp, body=None):
    msg = str(ts) + method + '/v2/' + endp
    if body:
        msg += json.dumps(body)
    return hmac.new(WARM_API_SECRET.encode('utf-8'), msg.encode(), hashlib.sha256).hexdigest()

def warm_exchange_req(method, endpoint, params=None):
    ts = int(time.time() * 1000)
    headers = {
        f'{Warm_API_Name}-Access-Key': WARM_API_KEY,
        f'{Warm_API_Name}-Access-Timestamp': str(ts),
        f'{Warm_API_Name}-Access-Signature': create_signature(ts, method, endpoint, params)
    }
    try:
        resp = requests.request(method, WARM_API_URL + endpoint, headers=headers, params=params)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as e:
        logging.error(f"API error ({endpoint}): {e}")
        return None

def get_warm_exchange_ticker(market):
    return warm_exchange_req('GET', f"ticker/price?market={market}")

def get_warm_exchange_balance():
    data = warm_exchange_req('GET', "balance")
    if data:
        return {item['symbol']: {'available': float(item['available']), 'in_order': float(item['inOrder'])} for item in data}
    return {}

def get_crypto_ticker(crypto):
    if crypto == "EUR":
        return {'eur_rate': 1.0, 'updated': int(time.time())}
    data = get_warm_exchange_ticker(f"{crypto}-EUR")
    if data and 'price' in data:
        return {'eur_rate': float(data['price']), 'updated': int(time.time())}
    logging.warning(f"Price not found for {crypto}")
    return None


def is_excel_installed():
    """
    Checks if Microsoft Excel is installed and accessible via COM.
    Returns True if Excel can be launched, False otherwise.
    """
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Quit()
        return True
    except Exception as e:
        logging.warning(f"Excel COM not available: {e}")
        return False

def read_cold_storage_xlsx(file_path='tracker.xlsx'):
    try:
        df = pd.read_excel(file_path, sheet_name='Cold_Storage', header=1)
        return dict(zip(df['Coin'].str.upper(), df['Amount']))
    except Exception as e:
        logging.error(f"Error reading cold storage from '{file_path}': {e}")
        return {}

# Additional functions (data processing, GUI, Excel export, etc.) continue here...
def calculate_buy_stake_sell_data(csv_filepath, real_time_prices):
    data_per_coin = defaultdict(lambda: {
        'bought_amount': 0,
        'bought_cost': 0,
        'staked_amount': 0,
        'deposited_amount': 0,
        'withdrawn_amount': 0,
        'sold_amount': 0,
        'sold_revenue': 0,
        'total_fees_eur': 0,
        'total_fees_crypto': 0
    })
    calc_message=""
    calc_message = "Calculating Fields...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()
    finished_button.config(state=tk.DISABLED)


    eur_invested = 0.0
    eur_withdrawn = 0.0
    all_rows = []

    try:
        with open(csv_filepath, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                all_rows.append(row)

                tx_type = row.get('Type', '').lower().strip()
                currency = row.get('Currency', '').upper().strip()
                amount = float(row.get('Amount', '0') or 0)
                quote_currency = row.get('Quote Currency', '').upper().strip()
                quote_price = float(row.get('Quote Price', '0') or 0)
                received_paid_currency = row.get('Received / Paid Currency', '').upper().strip()
                received_paid_amount = float(row.get('Received / Paid Amount', '0') or 0)
                fee_currency = row.get('Fee currency', '').upper().strip()
                fee_amount = float(row.get('Fee amount', '0') or 0)

                if amount == 0:
                    continue

                if currency == 'EUR':
                    if tx_type == 'deposit':
                        eur_invested += abs(amount)
                    elif tx_type == 'withdrawal':
                        eur_withdrawn += abs(amount)
                    continue

                if currency not in real_time_prices:
                    continue

                if tx_type == 'buy' and quote_currency == 'EUR':
                    cost = amount * quote_price
                    data_per_coin[currency]['bought_amount'] += amount
                    data_per_coin[currency]['bought_cost'] += cost

                elif tx_type == 'sell' and quote_currency == 'EUR':
                    revenue = abs(amount) * quote_price
                    data_per_coin[currency]['sold_amount'] += abs(amount)
                    data_per_coin[currency]['sold_revenue'] += revenue

                elif tx_type == 'staking':
                    data_per_coin[currency]['staked_amount'] += amount

                elif tx_type == 'deposit' and currency != 'EUR':
                    data_per_coin[currency]['deposited_amount'] += amount

                elif tx_type == 'withdrawal' and currency != 'EUR':
                    data_per_coin[currency]['withdrawn_amount'] += abs(amount)

                if fee_amount > 0 and fee_currency:
                    if fee_currency == 'EUR':
                        data_per_coin[currency]['total_fees_eur'] += fee_amount
                    elif fee_currency in real_time_prices:
                        fee_value_eur = fee_amount * real_time_prices[fee_currency]
                        data_per_coin[fee_currency]['total_fees_eur'] += fee_value_eur
                    else:
                        data_per_coin[fee_currency]['total_fees_crypto'] += fee_amount

        result = {}
        for coin, data in sorted(data_per_coin.items()):
            warm_balance = (
                data['bought_amount'] +
                data['staked_amount'] +
                data['deposited_amount'] -
                data['withdrawn_amount'] -
                data['sold_amount']
            )

            avg_buy_price = data['bought_cost'] / data['bought_amount'] if data['bought_amount'] > 0 else 0
            avg_sell_price = data['sold_revenue'] / data['sold_amount'] if data['sold_amount'] > 0 else 0
            current_price = real_time_prices.get(coin, 0)
            current_warm_value = warm_balance * current_price
            total_invested_this_coin = data['bought_cost'] - data['sold_revenue']

            if current_warm_value < 1 and warm_balance < 0.0001:
                continue

            result[coin] = {
                'bought': round(data['bought_amount'], 8),
                'staked': round(data['staked_amount'], 8),
                'deposited': round(data['deposited_amount'], 8),
                'withdrawn': round(data['withdrawn_amount'], 8),
                'sold': round(data['sold_amount'], 8),
                'warm_balance': round(warm_balance, 8),
                'avg_buy_price': round(avg_buy_price, 4),
                'avg_sell_price': round(avg_sell_price, 4),
                'current_price': round(current_price, 4),
                'current_warm_value': round(current_warm_value, 2),
                'total_invested': round(total_invested_this_coin, 2),
                'sold_revenue': round(data['sold_revenue'], 2),
                'fees_eur': round(data['total_fees_eur'], 2),
                'fees_crypto': round(data['total_fees_crypto'], 8)
            }
        calc_message=""
        calc_message = "Fished Calculating...."
        calc_message_label.config(text=calc_message)
        calc_message_label.update_idletasks()
        finished_button.config(state=tk.DISABLED)

        return result, round(eur_invested, 2), round(eur_withdrawn, 2), all_rows

    except Exception as e:
        logging.error(f"Error reading CSV: {e}")
        return {}, 0.0, 0.0, []

def exit_program():
    root.destroy()



def load_app_settings():
    """Load configuration from tracker.cfg into a flat dictionary."""
    config = configparser.ConfigParser()
    app_settings = {
        'refresh_main': 3,
        'refresh_warm': 3,
        'refresh_cold': 3,
        'refresh_total': 3,
        'write_warm': False,
        'write_cold': False,
        'write_total': True,
        'write_csv': True,
        'url1': '', 'Name1': '',
        'url2': '', 'Name2': '',
        'url3': '', 'Name3': '',
        'debug_mode': False,
        'dark_mode': True,
        'cold_storage_available': False
    }

    try:
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)

            if config.has_section('RefreshRate'):
                app_settings['refresh_main'] = config.getint('RefreshRate', 'Main', fallback=3)
                app_settings['refresh_warm'] = config.getint('RefreshRate', 'Warm', fallback=3)
                app_settings['refresh_cold'] = config.getint('RefreshRate', 'Cold', fallback=3)
                app_settings['refresh_total'] = config.getint('RefreshRate', 'Total', fallback=3)

            if config.has_section('WriteData'):
                app_settings['write_warm'] = config.getboolean('WriteData', 'Warm', fallback=False)
                app_settings['write_cold'] = config.getboolean('WriteData', 'Cold', fallback=False)
                app_settings['write_total'] = config.getboolean('WriteData', 'Total', fallback=True)
                app_settings['write_csv'] = config.getboolean('WriteData', 'CSV', fallback=True)

            if config.has_section('OptionalURL'):
                for i in range(1, 4):
                    app_settings[f'url{i}'] = config.get('OptionalURL', f'URL{i}', fallback='')
                    app_settings[f'Name{i}'] = config.get('OptionalURL', f'Name{i}', fallback='')

            if config.has_section('Miscellaneous'):
                app_settings['debug_mode'] = config.getboolean('Miscellaneous', 'DebugMode', fallback=False)
                app_settings['dark_mode'] = config.getboolean('Miscellaneous', 'darkmod', fallback=True)
                app_settings['cold_storage_available'] = config.getboolean('Miscellaneous', 'Cold Storage Available', fallback=False)

    except Exception as e:
        logging.warning(f"Error loading configuration: {e}")
        logging.info("Using default settings...")

    return app_settings

def write_treeview_to_excel(filename, sheet_name, treeview_widget, columns_tuple,
                            eur_invested, eur_withdrawn, net_invested,
                            total_warm_value, total_cold_value, current_value):
    """
    Writes the contents of a Tkinter Treeview to an Excel sheet.
    Includes summary values like EUR in/out, total value, and P/L.
    """
    calc_message=""
    calc_message = "Creating Treeview...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()
    finished_button.config(state=tk.DISABLED)

    try:
        try:
            book = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"File '{filename}' not found.")
            return

        # Delete existing sheet if it exists
        if sheet_name in book.sheetnames:
            del book[sheet_name]

        ws = book.create_sheet(sheet_name)

        # Write timestamp
        current_datetime_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([f"Data as shown on screen, generated on: {current_datetime_str}"])

        # Write column headers
        ws.append(list(columns_tuple))

        # Write Treeview data
        for item_id in treeview_widget.get_children():
            row_values = treeview_widget.item(item_id, 'values')
            ws.append([str(val) for val in row_values])

        # Adjust column widths
        for col_idx, column_title in enumerate(columns_tuple, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column_title))
            for row_idx in range(3, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

        # Write summary section
        last_row = ws.max_row
        ws.cell(row=1, column=12, value="Invested")
        ws.cell(row=2, column=12, value=eur_invested)
        ws.cell(row=1, column=13, value="Withdrawn")
        ws.cell(row=2, column=13, value=eur_withdrawn)
        ws.cell(row=1, column=14, value="Net Invest")
        ws.cell(row=2, column=14, value=net_invested)
        ws.cell(row=1, column=15, value="Warm Value")
        ws.cell(row=2, column=15, value=total_warm_value)
        ws.cell(row=1, column=16, value="Cold Value")
        ws.cell(row=2, column=16, value=round(total_cold_value, 2))
        ws.cell(row=1, column=17, value="Total Value")
        ws.cell(row=2, column=17, value=round(current_value, 2))

        # Profit/Loss
        ws.cell(row=4, column=12, value="P/L Value")
        pl_value = current_value - net_invested
        ws.cell(row=5, column=12, value=round(pl_value, 2))
        ws.cell(row=4, column=13, value="P/L %")
        pl_percentage = ((pl_value / net_invested)/net_invested) * 100 if net_invested != 0 else 0
        ws.cell(row=5, column=13, value=round(pl_percentage, 2))

        book.save(filename)
        calc_message=""
        calc_message = "Treeview Data Written...."
        calc_message_label.config(text=calc_message)
        calc_message_label.update_idletasks()


        logging.info(f"Treeview data successfully written to '{filename}' in sheet '{sheet_name}'.")

    except PermissionError:
        messagebox.showerror("Permission Error", f"Could not save '{filename}'. Is the file open?")
    except Exception as e:
        logging.error(f"Error writing Treeview to Excel: {e}")
        messagebox.showerror("Excel Error", f"Error writing to '{sheet_name}': {e}")

def exit_program():
    root.destroy()

def browse_file():
    calc_message=""
    calc_message = "Waiting for CSV File...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()
    finished_button.config(state=tk.DISABLED)
    filepath = filedialog.askopenfilename(
        title="Select Bitvavo CSV file",
        filetypes=[("CSV files", "*.csv")],
        initialdir=os.getcwd()
    )


    if not filepath:
        return

    cold_storage_data = read_cold_storage_xlsx()
    warm_storage_data = get_warm_exchange_balance()
    relevant_coins = list(set(cold_storage_data.keys()).union(set(warm_storage_data.keys())))

    real_time_prices = {}
    for coin in relevant_coins:
        if coin == 'EUR':
            real_time_prices[coin] = 1.0
        else:
            ticker_data = get_warm_exchange_ticker(f"{coin}-EUR")
            if ticker_data and 'price' in ticker_data:
                real_time_prices[coin] = float(ticker_data['price'])
            else:
                real_time_prices[coin] = 0
                logging.warning(f"Could not fetch price for {coin}. Defaulting to 0.")

    results, eur_invested, eur_withdrawn, all_csv_data = calculate_buy_stake_sell_data(filepath, real_time_prices)

    # Clear existing rows
    for row in treeview.get_children():
        treeview.delete(row)

    total_warm_value = 0
    total_cold_value = 0
    total_invested_all = 0

    if results:
        for coin, data in results.items():
            cold_amount = cold_storage_data.get(coin, 0)
            cold_value = cold_amount * data['current_price']

            if data['warm_balance'] >= 0.00000001 or cold_amount >= 0.00000001:
                total_amount = data['warm_balance'] + cold_amount
                treeview.insert('', 'end', values=(
                    coin,
                    f"{data['warm_balance']:.4f}",
                    f"{cold_amount:.4f}",
                    f"{total_amount:.2f}",
                    f"€{data['avg_buy_price']:.4f}" if data['avg_buy_price'] > 0 else "N/A",
                    f"€{data['avg_sell_price']:.4f}" if data['avg_sell_price'] > 0 else "N/A",
                    f"€{data['current_warm_value']:.2f}",
                    f"€{cold_value:.2f}",
                    f"€{data['total_invested']:.2f}",
                    f"€{data['fees_eur']:.2f}"
                ))

                total_warm_value += data['current_warm_value']
                total_cold_value += cold_value
                total_invested_all += data['total_invested']

        net_invested = eur_invested - eur_withdrawn
        eur_in_label.config(text=f"Total EUR In: €{eur_invested:.2f}")
        eur_out_label.config(text=f"Total EUR Out: €{eur_withdrawn:.2f}")
        total_invest_label.config(text=f"Net Invested: €{net_invested:.2f}")
        warm_value_label.config(text=f"Warm Value: €{total_warm_value:.2f}")
        cold_value_label.config(text=f"Cold Value: €{total_cold_value:.2f}")
        total_value_label.config(text=f"Total Value: €{(total_warm_value + total_cold_value):.2f}")
        finished_button.config(state=tk.NORMAL)

        current_value = total_warm_value + total_cold_value
        if app_settings.get('write_csv'):
            write_treeview_to_excel("tracker.xlsx", "CSV_History", treeview, columns,
                                    eur_invested, eur_withdrawn, net_invested,
                                    total_warm_value, total_cold_value, current_value)
            create_excel_with_pivots("tracker.xlsx", all_csv_data, relevant_coins if results else [])
    else:
        treeview.insert('', 'end', values=("No relevant transactions found.",))
        for label in [eur_in_label, eur_out_label, total_invest_label, warm_value_label, cold_value_label, total_value_label]:
            label.config(text="€0.00")
        finished_button.config(state=tk.DISABLED)

def flatten_cols(cols):
    """Flattens multi-level column headers into a single list of strings."""
    new_cols = []
    for col in cols:
        if isinstance(col, tuple):
            new_cols.append(' '.join(str(c) for c in col if c))
        else:
            new_cols.append(str(col))
    return new_cols

def create_excel_with_pivots(filename, all_data, relevant_coins):
    calc_message = "Building Pivot Tables ...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()

    """
    Creates pivot tables in Excel.
    On Windows: uses Excel COM automation to create real PivotTables if Excel is installed.
    On other systems or failure: generates static pivot tables using pandas and openpyxl.
    """
    if platform.system() == "Windows" and is_excel_installed():
        try:
            create_real_pivot_excel_windows(filename, all_data, relevant_coins)
            return
        except Exception as e:
            logging.warning(f"Failed to create real PivotTables. Falling back to static version: {e}")

    # Fallback for non-Windows or COM failure
    create_static_pivot_excel(filename, all_data, relevant_coins)

def create_real_pivot_excel_windows(filename, all_data, relevant_coins):
    """
    Creates two real Excel PivotTables directly inside the existing tracker.xlsx file using COM automation.
    Only works on Windows with Excel installed.
    """
    import win32com.client
    calc_message=""
    calc_message = "Creating Real Pivot Tables...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()

    df = pd.DataFrame(all_data)
    df['Currency'] = df['Currency'].str.upper()
    df_filtered = df[df['Currency'].isin(relevant_coins) | (df['Currency'] == 'EUR')]

    excel = win32com.client.Dispatch("Excel.Application")
    excel.DisplayAlerts = False
    excel.Visible = False
    calc_message=""
    calc_message = "Creating Real Pivot Tables...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()



    try:
        wb = excel.Workbooks.Open(os.path.abspath(filename))

        # Remove old sheets if they exist
        for sheet_name in ["Raw Data", "Pivot Table Summary", "Pivot Table Detailed"]:
            try:
                wb.Sheets(sheet_name).Delete()
            except Exception:
                pass

        # Add Raw Data sheet
        raw_ws = wb.Sheets.Add()
        raw_ws.Name = "Raw Data"

        for col_idx, header in enumerate(df_filtered.columns, start=1):
            raw_ws.Cells(1, col_idx).Value = header

        for row_idx, row in enumerate(df_filtered.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                raw_ws.Cells(row_idx, col_idx).Value = value

        last_row = df_filtered.shape[0] + 1
        last_col = df_filtered.shape[1]
        data_range = raw_ws.Range(raw_ws.Cells(1, 1), raw_ws.Cells(last_row, last_col))

        # Pivot Table Summary
        pivot_ws_summary = wb.Sheets.Add()
        pivot_ws_summary.Name = "Pivot Table Summary"

        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)
        pivot_table_summary = pivot_cache.CreatePivotTable(
            TableDestination=pivot_ws_summary.Cells(3, 2),
            TableName="CryptoPivotSummary"
        )

        pivot_table_summary.PivotFields("Currency").Orientation = 1  # xlRowField
        pivot_table_summary.PivotFields("Type").Orientation = 2      # xlColumnField
        pivot_table_summary.AddDataField(pivot_table_summary.PivotFields("Amount"), "Sum of Amount", -4157)
        pivot_table_summary.AddDataField(pivot_table_summary.PivotFields("Fee amount"), "Sum of Fees", -4157)
        pivot_table_summary.AddDataField(pivot_table_summary.PivotFields("Received / Paid Amount"), "Sum of EUR", -4157)

        # Pivot Table Detailed
        pivot_ws_detailed = wb.Sheets.Add()
        pivot_ws_detailed.Name = "Pivot Table Detailed"

        pivot_cache2 = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)
        pivot_table_detailed = pivot_cache2.CreatePivotTable(
            TableDestination=pivot_ws_detailed.Cells(3, 2),
            TableName="CryptoPivotDetailed"
        )

        for field in ["Timezone", "Date", "Time", "Type", "Currency"]:
            pivot_table_detailed.PivotFields(field).Orientation = 1  # xlRowField

        for field in ["Quote Currency", "Received / Paid Currency", "Fee currency"]:
            pivot_table_detailed.PivotFields(field).Orientation = 2  # xlColumnField

        pivot_table_detailed.AddDataField(pivot_table_detailed.PivotFields("Amount"), "Sum of Amount", -4157)
        pivot_table_detailed.AddDataField(pivot_table_detailed.PivotFields("Quote Price"), "Avg Quote Price", -4106)
        pivot_table_detailed.AddDataField(pivot_table_detailed.PivotFields("Fee amount"), "Sum of Fees", -4157)
        pivot_table_detailed.AddDataField(pivot_table_detailed.PivotFields("Received / Paid Amount"), "Sum of EUR", -4157)

        wb.Save()
        wb.Close(SaveChanges=True)
        calc_message=""
        calc_message = "Finished Creating!"
        calc_message_label.config(text=calc_message)
        calc_message_label.update_idletasks()

        finished_button.config(state="normal")

        logging.info(f"Real PivotTables created in '{filename}' using Excel COM.")

    except Exception as e:
        logging.error(f"Failed to create real PivotTables: {e}")
        raise RuntimeError(f"Failed to create real PivotTables: {e}")

    finally:
        try:
            excel.Quit()
        except Exception:
            pass

def create_static_pivot_excel(filename, all_data, relevant_coins):
    """
    Creates static pivot tables using pandas and openpyxl.
    This version works on all platforms.
    """
    calc_message=""
    calc_message = "Creating Static Pivot Tables...."
    calc_message_label.config(text=calc_message)
    calc_message_label.update_idletasks()
    try:
        if os.path.exists(filename):
            book = openpyxl.load_workbook(filename)
            for sheet_name in ['Raw Data', 'Pivot Table Summary', 'Pivot Table Detailed']:
                if sheet_name in book.sheetnames:
                    del book[sheet_name]
        else:
            book = openpyxl.Workbook()

        df_raw = pd.DataFrame(all_data)
        ws_raw = book.create_sheet('Raw Data')
        ws_raw.append(list(df_raw.columns))
        for row in df_raw.itertuples(index=False):
            ws_raw.append(list(row))

        for col in ['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount']:
            if col in df_raw.columns:
                df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce')

        df_raw['Currency'] = df_raw['Currency'].str.upper()
        df_filtered = df_raw[df_raw['Currency'].isin(relevant_coins) | (df_raw['Currency'] == 'EUR')]

        if not df_filtered.empty:
            # Summary Pivot
            pivot_summary = pd.pivot_table(
                df_filtered,
                values=['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount'],
                index='Currency',
                columns='Type',
                aggfunc='sum',
                margins=True,
                margins_name='Total'
            )
            ws_summary = book.create_sheet('Pivot Table Summary')
            flat_cols = [pivot_summary.index.name] + flatten_cols(pivot_summary.columns)
            ws_summary.append(flat_cols)
            for index, row in pivot_summary.iterrows():
                ws_summary.append([index] + list(row))

            # Detailed Pivot
            pivot_detailed = pd.pivot_table(
                df_filtered,
                values=['Amount', 'Quote Price', 'Fee amount', 'Received / Paid Amount'],
                index=['Timezone', 'Date', 'Time', 'Type', 'Currency'],
                columns=['Quote Currency', 'Received / Paid Currency', 'Fee currency'],
                aggfunc='sum',
                fill_value=0
            )
            ws_detailed = book.create_sheet('Pivot Table Detailed')
            flat_cols = list(pivot_detailed.index.names) + flatten_cols(pivot_detailed.columns)
            ws_detailed.append(flat_cols)
            for index, row in pivot_detailed.iterrows():
                ws_detailed.append(list(index) + list(row))

        book.save(filename)
        calc_message=""
        calc_message = "Finished Creating!"
        calc_message_label.config(text=calc_message)
        calc_message_label.update_idletasks()

        finished_button.config(state="normal")

        logging.info(f"Excel file '{filename}' updated with static pivot tables.")

    except Exception as e:
        logging.error(f"Error creating static pivot tables: {e}")
        messagebox.showerror("Excel Error", f"Error creating pivot tables: {e}")



# Initialize GUI
root = tk.Tk()
root.title("Bitvavo CSV Analyzer - Crypto Price Tracker V1.5")
root.geometry("1100x650")
icon_path = os.path.join(os.getcwd(), "crypto", f"calc.ico")
root.iconbitmap(icon_path)
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

tk.Label(root, text="Select Bitvavo CSV File", font=("Helvetica", 14)).pack(pady=10)
tk.Button(root, text="Browse...", command=browse_file, width=20).pack(pady=5)

# Load settings
app_settings = load_app_settings()
par_write_csv = app_settings.get('write_csv', True)

# Define Treeview columns
columns = ('Coin', 'Warm Amount', 'Cold Amount', 'Total', 'Avg Buy €', 'Avg Sell €',
           'Warm Value €', 'Cold Value €', 'Invested €', 'Fees €')

treeview = ttk.Treeview(root, columns=columns, show='headings')

for col in columns:
    treeview.heading(col, text=col)
    if col in ['Warm Amount', 'Cold Amount']:
        treeview.column(col, minwidth=120, width=120, stretch=True)
    elif 'Value' in col or 'Invested' in col or 'Fees' in col:
        treeview.column(col, minwidth=100, width=100, stretch=True)
    else:
        treeview.column(col, minwidth=80, width=100, stretch=True)

treeview.pack(padx=10, pady=10, fill="both", expand=True)

# Scrollbars
scrollbar_y = tk.Scrollbar(root, orient="vertical", command=treeview.yview)
scrollbar_y.pack(side="right", fill="y")
treeview.configure(yscrollcommand=scrollbar_y.set)

scrollbar_x = tk.Scrollbar(root, orient="horizontal", command=treeview.xview)
scrollbar_x.pack(side="bottom", fill="x")
treeview.configure(xscrollcommand=scrollbar_x.set)

# Totals frame
totals_frame = tk.Frame(root)
totals_frame.pack(pady=10)

# First row of labels
first_row = tk.Frame(totals_frame)
first_row.pack()

eur_in_label = tk.Label(first_row, text="Total EUR In: €0.00", font=("Helvetica", 10))
eur_in_label.pack(side="left", padx=10)

eur_out_label = tk.Label(first_row, text="Total EUR Out: €0.00", font=("Helvetica", 10))
eur_out_label.pack(side="left", padx=10)

total_invest_label = tk.Label(first_row, text="Net Invested: €0.00", font=("Helvetica", 10))
total_invest_label.pack(side="left", padx=10)

# Second row of labels
second_row = tk.Frame(totals_frame)
second_row.pack()



warm_value_label = tk.Label(second_row, text="Warm Value: €0.00", font=("Helvetica", 10), fg="orange")
warm_value_label.pack(side="left", padx=10)

cold_value_label = tk.Label(second_row, text="Cold Value: €0.00", font=("Helvetica", 10), fg="blue")
cold_value_label.pack(side="left", padx=10)

total_value_label = tk.Label(second_row, text="Total Value: €0.00", font=("Helvetica", 12, "bold"), fg="green")
total_value_label.pack(side="left", padx=10)


# bottom line  calc_message and  Finished button
bottom_row = tk.Frame(root)
bottom_row.pack(fill="x", pady=10, padx=10)

# calc_message label helemaal links
calc_message_label = tk.Label(bottom_row, text="", font=("Helvetica", 10, "italic","bold"), fg="black")
calc_message_label.pack(side="left")

# Finished button helemaal rechts
finished_button = tk.Button(bottom_row, text="Finished", command=exit_program,
                            bg="lightgray", fg="forestgreen", font=("Helvetica", 12), state=tk.DISABLED)

finished_button.pack(side="right")

# Start GUI loop
calc_message = "Initializing...."
calc_message_label.config(text=calc_message)
root.mainloop()
